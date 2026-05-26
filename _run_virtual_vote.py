"""
가상 여론조사 러너 — Nemotron 페르소나 × 2026 지방선거

입력:
  - pledges_text.json  {huboId: text}  (사전 추출됨)
  - pledges.json       {huboId: {giho, name, party, sggname, ...}}
  - nec_2026_지방선거_후보자.xlsx (race 목록)
  - Nemotron-Personas-Korea/by_province/*.csv (페르소나 풀)

처리:
  1. 시도지사·구시군의장·교육감 race 목록 추출
  2. 각 race 별로 후보 익명화 (기호 오름차순 → A/B/C…)
  3. race region 매칭 페르소나 N명 샘플링 (seed=42)
  4. 각 페르소나에 대해 Gemini Flash Lite 호출 → JSON 응답
  5. 매 race 완료 시 virtual_poll_raw.jsonl append (재실행 시 이어서)
  6. 모두 끝나면 virtual_poll.json 으로 집계

사용:
  GEMINI_API_KEY=... python3 _run_virtual_vote.py
  GEMINI_API_KEY=... python3 _run_virtual_vote.py --personas-per-race 20 --only "시도지사선거|서울특별시"
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import os
import random
import re
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import requests

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("virtual_vote")

ROOT = Path(__file__).parent
PERSONA_DIR = ROOT / "Nemotron-Personas-Korea" / "by_province"
XLSX = ROOT / "nec_2026_지방선거_후보자.xlsx"
PLEDGES_JSON = ROOT / "pledges.json"
PLEDGES_TEXT = ROOT / "pledges_text.json"
RAW_OUT = ROOT / "virtual_poll_raw.jsonl"
AGG_OUT = ROOT / "virtual_poll.json"

MODEL = "gemini-2.5-flash-lite"
ENDPOINT = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL}:generateContent"

# NEC 시도명 ↔ persona province 매핑
SIDO_TO_PROVS = {
    "서울특별시": ["서울"],
    "부산광역시": ["부산"],
    "대구광역시": ["대구"],
    "인천광역시": ["인천"],
    "광주광역시": ["광주"],
    "대전광역시": ["대전"],
    "울산광역시": ["울산"],
    "세종특별자치시": ["세종"],
    "경기도": ["경기"],
    "강원특별자치도": ["강원"],
    "충청북도": ["충청북"],
    "충청남도": ["충청남"],
    "전북특별자치도": ["전북"],
    "전라남도": ["전라남"],
    "경상북도": ["경상북"],
    "경상남도": ["경상남"],
    "제주특별자치도": ["제주"],
    # 시도지사/교육감용 통합 권역
    "전남광주통합특별시": ["광주", "전라남"],
}


# ---------- persona pool ---------- #

class PersonaPool:
    def __init__(self):
        self._by_prov: dict[str, list[dict]] = {}

    def load(self, provinces: set[str]):
        for prov in provinces:
            if prov in self._by_prov:
                continue
            path = PERSONA_DIR / f"{prov}.csv"
            if not path.exists():
                log.warning("province csv 없음: %s", path)
                self._by_prov[prov] = []
                continue
            rows = []
            with open(path, encoding="utf-8-sig") as f:
                r = csv.DictReader(f)
                for row in r:
                    rows.append(row)
            self._by_prov[prov] = rows
            log.info("loaded persona %s: %d", prov, len(rows))

    def sample_for_sido(self, sido_name: str, n: int, rng: random.Random) -> list[dict]:
        provs = SIDO_TO_PROVS.get(sido_name)
        if not provs:
            return []
        self.load(set(provs))
        pool = []
        for p in provs:
            pool.extend(self._by_prov.get(p, []))
        if not pool:
            return []
        if len(pool) <= n:
            return list(pool)
        return rng.sample(pool, n)

    def sample_for_sgg(self, sido_name: str, sgg_name: str, n: int, rng: random.Random) -> list[dict]:
        # 구시군장: persona district 의 시군구 부분 == sgg_name
        provs = SIDO_TO_PROVS.get(sido_name, [])
        if not provs:
            return []
        self.load(set(provs))
        pool = []
        for p in provs:
            for row in self._by_prov.get(p, []):
                d = row.get("district") or ""
                if "-" not in d:
                    continue
                sgg_part = d.split("-", 1)[1].strip()
                if sgg_part == sgg_name:
                    pool.append(row)
        if not pool:
            return []
        if len(pool) <= n:
            return list(pool)
        return rng.sample(pool, n)


# ---------- race extraction ---------- #

def extract_races(pledges: dict, pledges_text: dict) -> list[dict]:
    """xlsx 에서 시도지사·구시군의장·교육감 race 목록을 빌드. 각 race 마다 후보 리스트 포함."""
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    races: dict[str, dict] = {}

    sheet_configs = [
        ("시도지사선거", "sido"),
        ("구시군의장선거", "sgg"),
        ("교육감선거", "sido"),
    ]
    for sheet_name, level in sheet_configs:
        if sheet_name not in wb.sheetnames:
            log.warning("xlsx 시트 없음: %s", sheet_name)
            continue
        ws = wb[sheet_name]
        hdrs = [c.value for c in next(ws.iter_rows(max_row=1))]
        idx = {h: i for i, h in enumerate(hdrs)}

        for row in ws.iter_rows(min_row=2, values_only=True):
            sido = row[idx["시도"]]
            sgg = row[idx["선거구"]]
            giho = row[idx["기호"]]
            party = row[idx["정당"]]
            name = row[idx["성명"]]
            hubo_id = str(row[idx["huboId"]]).strip()
            if not hubo_id:
                continue

            # race key
            if level == "sido":
                region = sido
                key = f"{sheet_name}|{sido}"
            else:
                region = sgg or sido
                key = f"{sheet_name}|{sido}|{region}"

            if key not in races:
                races[key] = {
                    "sheet": sheet_name,
                    "sido": sido,
                    "region": region,    # display
                    "level": level,
                    "candidates": [],
                }
            races[key]["candidates"].append({
                "huboId": hubo_id,
                "giho": str(giho) if giho is not None else "",
                "name": name,
                "party": _normalize_party(party),
                "has_pledge": bool(pledges_text.get(hubo_id)),
            })

    # 후보 정렬 (기호 오름차순) + 알파벳 부여
    out = []
    for key, race in races.items():
        cands = sorted(race["candidates"], key=lambda c: _giho_int(c["giho"]))
        for i, c in enumerate(cands):
            c["letter"] = chr(ord("A") + i)
        race["candidates"] = cands
        race["key"] = key
        out.append(race)
    log.info("races: %d (시도지사·구시군장·교육감)", len(out))
    return out


def _normalize_party(p):
    if not p: return "무소속"
    return re.sub(r"\(\d+\)\s*$", "", str(p)).strip() or "무소속"


def _giho_int(giho):
    try: return int(re.sub(r"[^0-9]", "", str(giho)) or "99")
    except: return 99


# ---------- prompt builder ---------- #

def build_voter_prompt(persona: dict, race: dict, anonymized_pledges: list[tuple[str, str]]) -> str:
    cand_block = "\n\n".join(
        f"후보 {letter}\n{text}" for letter, text in anonymized_pledges
    )

    province = persona.get("province", "")
    district = persona.get("district", "")
    age = persona.get("age", "")
    sex = persona.get("sex", "")
    housing = persona.get("housing_type", "")
    family = persona.get("family_type", "")
    edu = persona.get("education_level", "")
    job = persona.get("occupation", "")

    persona_summary = (persona.get("persona") or "")[:240]
    hobbies = (persona.get("hobbies_and_interests") or "")[:200]
    cultural = (persona.get("cultural_background") or "")[:180]

    valid_letters = "/".join(letter for letter, _ in anonymized_pledges)

    return f"""당신은 대한민국 {province} {district}에 거주하는 유권자입니다. 아래 페르소나에 일관되게 답해 주세요.

[페르소나]
{persona_summary}
나이·성별: {age}세 {sex}
거주·가족: {housing}, {family}
학력·직업: {edu} · {job}
관심사: {hobbies}
배경: {cultural}

[지침]
1) 페르소나의 경제적 처지·생활 반경·가족 구성·가치관이 1차 기준입니다. 추상적 이념보다 "내 삶에 무엇이 바뀌는가"가 우선.
2) 모두 마음에 안 들면 "지지없음(none)" 또는 "기권(abstain)" 가능. 한국 지방선거 평균 투표율 ~50%, 무당파 30~40%.
3) 후보·정당은 알파벳으로 익명화. 정당 색채로 추측하지 말고 공약 내용으로만 판단.
4) 광역장·기초장·교육감은 같은 날 함께 투표하지만, 지금은 {race["sheet"]} 만 판단하세요.
5) 출력은 지정된 JSON 만. 다른 설명·서두·마크다운 금지.

[선거]
지역: {race["sido"]} {race["region"] if race["level"] == "sgg" else ""}
선거: {race["sheet"]}

[후보 공약 — 익명]
{cand_block}

[질문]
① 각 후보 공약이 당신 일상에 줄 영향 1~5점, 마음에 드는 공약 1개, 거부감 드는 공약 1개
② 각 후보 호감도 1~10점
③ 지지 후보 / 강도 / 결정에 가장 큰 영향을 준 분야
④ 한 줄 코멘트 (페르소나 말투)

[응답 JSON 형식 — 이것 외에 다른 텍스트 절대 금지]
{{
  "ratings": {{
    "A": {{"impact":1~5,"likeability":1~10,"best":"...","concern":"..."}},
    "B": {{...}}
  }},
  "vote": "{valid_letters}|none|abstain",
  "confidence": "strong|moderate|reluctant",
  "key_issue": "...",
  "comment": "..."
}}"""


# ---------- gemini call ---------- #

def call_gemini(api_key: str, prompt: str, timeout: int = 60) -> dict | None:
    body = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0.7,
            "maxOutputTokens": 800,
            "responseMimeType": "application/json",
        },
    }
    url = f"{ENDPOINT}?key={api_key}"
    last_err = None
    for attempt in range(3):
        try:
            r = requests.post(url, json=body, timeout=timeout)
            if r.status_code == 429:
                # rate limit — backoff
                time.sleep(2 ** attempt)
                continue
            if not r.ok:
                last_err = f"HTTP {r.status_code}: {r.text[:200]}"
                time.sleep(1)
                continue
            j = r.json()
            parts = j.get("candidates", [{}])[0].get("content", {}).get("parts", [])
            text = "".join(p.get("text", "") for p in parts).strip()
            return _parse_json_loose(text)
        except Exception as e:
            last_err = str(e)
            time.sleep(1)
    log.debug("gemini fail: %s", last_err)
    return None


def _parse_json_loose(text: str) -> dict | None:
    if not text: return None
    # strip markdown fences
    text = re.sub(r"^```(?:json)?\s*", "", text.strip())
    text = re.sub(r"\s*```\s*$", "", text)
    # find first {...}
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m: return None
    try:
        return json.loads(m.group(0))
    except json.JSONDecodeError:
        return None


# ---------- runner ---------- #

def already_done_keys(raw_path: Path) -> set[str]:
    if not raw_path.exists(): return set()
    done = set()
    with raw_path.open(encoding="utf-8") as f:
        for line in f:
            try:
                r = json.loads(line)
                done.add(r["race_key"])
            except: pass
    return done


def run_race(api_key, race, pledges_text, persona_pool, n_per_race, rng, workers):
    candidates = race["candidates"]
    valid = [c for c in candidates if c.get("has_pledge")]
    if len(valid) < 2:
        log.info("  [%s] 공약 제출 후보 부족 (%d) — skip", race["key"], len(valid))
        return None

    # anonymize: A..N for valid candidates (sorted by giho already)
    anon = []
    for i, c in enumerate(valid):
        letter = chr(ord("A") + i)
        c["letter"] = letter
        text = (pledges_text.get(c["huboId"]) or "").strip()
        if len(text) > 4500:
            text = text[:4500] + "\n…(이하 생략)"
        anon.append((letter, text))

    # sample personas
    if race["level"] == "sido":
        personas = persona_pool.sample_for_sido(race["sido"], n_per_race, rng)
    else:
        personas = persona_pool.sample_for_sgg(race["sido"], race["region"], n_per_race, rng)

    if not personas:
        log.warning("  [%s] persona 매칭 없음", race["key"])
        return None

    log.info("  [%s] 후보 %d, persona %d명 → voting", race["key"], len(valid), len(personas))

    # call Gemini in parallel
    def worker(persona):
        prompt = build_voter_prompt(persona, race, anon)
        result = call_gemini(api_key, prompt)
        return persona, result

    responses = []
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(worker, p) for p in personas]
        ok_count = 0
        for fut in as_completed(futs):
            persona, result = fut.result()
            if result is None:
                continue
            ok_count += 1
            responses.append({
                "persona_uuid": persona.get("uuid"),
                "persona_meta": {
                    "age": persona.get("age"),
                    "sex": persona.get("sex"),
                    "occupation": persona.get("occupation"),
                    "district": persona.get("district"),
                    "family_type": persona.get("family_type"),
                    "housing_type": persona.get("housing_type"),
                    "education_level": persona.get("education_level"),
                },
                "response": result,
            })
        log.info("  [%s] %d/%d 응답 OK", race["key"], ok_count, len(personas))

    return {
        "race_key": race["key"],
        "sheet": race["sheet"],
        "sido": race["sido"],
        "region": race["region"],
        "level": race["level"],
        "candidates": [
            {"huboId": c["huboId"], "giho": c["giho"], "name": c["name"],
             "party": c["party"], "letter": c.get("letter")}
            for c in candidates
        ],
        "anonymized_letters": [c["letter"] for c in valid],
        "n_personas_sampled": len(personas),
        "n_responses_ok": len(responses),
        "responses": responses,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--key", default=os.environ.get("GEMINI_API_KEY", ""))
    ap.add_argument("--personas-per-race", type=int, default=100)
    ap.add_argument("--workers", type=int, default=10)
    ap.add_argument("--only", help="race key prefix filter (e.g. '시도지사선거|서울특별시')")
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    if not args.key:
        log.error("GEMINI_API_KEY env 또는 --key 필요")
        sys.exit(1)

    # load inputs
    if not PLEDGES_JSON.exists():
        sys.exit("pledges.json 없음. 먼저 _fetch_pledges.py 실행")
    if not PLEDGES_TEXT.exists():
        sys.exit("pledges_text.json 없음. 먼저 _extract_pledge_text.py 실행")

    pledges = json.loads(PLEDGES_JSON.read_text(encoding="utf-8"))
    pledges_text = json.loads(PLEDGES_TEXT.read_text(encoding="utf-8"))
    log.info("pledges=%d, pledges_text=%d", len(pledges), len(pledges_text))

    races = extract_races(pledges, pledges_text)
    if args.only:
        races = [r for r in races if r["key"].startswith(args.only)]
        log.info("filtered by --only: %d races", len(races))

    done = already_done_keys(RAW_OUT)
    log.info("이미 완료된 race: %d (이어서 진행)", len(done))

    persona_pool = PersonaPool()
    rng = random.Random(args.seed)

    with RAW_OUT.open("a", encoding="utf-8") as f:
        for i, race in enumerate(races, 1):
            if race["key"] in done:
                continue
            log.info("[%d/%d] %s", i, len(races), race["key"])
            try:
                result = run_race(args.key, race, pledges_text, persona_pool,
                                  args.personas_per_race, rng, args.workers)
            except Exception as e:
                log.exception("race err: %s", e)
                continue
            if result is None:
                continue
            f.write(json.dumps(result, ensure_ascii=False) + "\n")
            f.flush()
            time.sleep(0.5)  # gentle pacing

    log.info("done. virtual_poll_raw.jsonl 에 누적됨. 집계는 build_virtual_poll.py 에서.")


if __name__ == "__main__":
    main()
