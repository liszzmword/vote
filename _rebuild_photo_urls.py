"""구시군 단위 선거 후보의 사진 URL 재구성.

문제: 크롤러가 photo URL 만들 때 시도 cityCode(1100=서울 등)를 썼는데
NEC의 사진 디렉토리는 구시군 단위 선거일 때 시군구 코드(1101=종로구 등)를 씀.

해결:
- 17개 시도 × electionCode=6 의 townCode JSON 받아 매핑 사전 구축
- 시도별 (구시군 이름 → 시군구 4자리 코드)
- 후보의 sgg(선거구) 필드에서 구시군명 추출 → 매핑 적용 → URL 재구성
- 시도 단위 선거(시도지사/광역비례/교육감)는 시도 cityCode 그대로
"""
from __future__ import annotations

import json
import re
import time
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).parent
XLSX = ROOT / "nec_2026_지방선거_후보자.xlsx"
OUT = ROOT / ".photo_urls.json"

BASE = "https://info.nec.go.kr"
ELECTION_ID = "0020260603"

# 시도명 → cityCode (사진 디렉토리에서 쓰는 값과 동일)
SIDO_TO_CODE: dict[str, str] = {
    "서울특별시": "1100", "부산광역시": "2600", "대구광역시": "2700",
    "인천광역시": "2800", "광주광역시": "2900", "대전광역시": "3000",
    "울산광역시": "3100", "세종특별자치시": "5100", "경기도": "4100",
    "강원특별자치도": "5200", "충청북도": "4300", "충청남도": "4400",
    "전북특별자치도": "5300", "전라남도": "4600", "경상북도": "4700",
    "경상남도": "4800", "제주특별자치도": "4900",
    "전남광주통합특별시": "2900",  # 통합시 사진은 광주(2900) 디렉토리에
}

# 시도 단위 선거 (sgg 없이 시도 cityCode 사용)
SIDO_LEVEL_ELECTIONS = {"시도지사선거", "광역의원비례대표선거", "교육감선거"}


def build_session() -> requests.Session:
    s = requests.Session()
    s.headers["User-Agent"] = "Mozilla/5.0"
    s.get(f"{BASE}/main/showDocument.xhtml",
          params={"electionId": ELECTION_ID, "topMenuId": "CP", "secondMenuId": "CPRI03"},
          timeout=30)
    return s


def fetch_town_map(session: requests.Session, city_code: str, election_code: str) -> dict[str, str]:
    """{구시군명: 코드}. townCode JSON. election_code=6=일반, 5=통합권역."""
    r = session.post(
        f"{BASE}/bizcommon/selectbox/selectbox_townCodeBySgJson.json",
        data={"electionId": ELECTION_ID, "electionCode": election_code, "cityCode": city_code},
        timeout=30,
    )
    body = r.json().get("jsonResult", {}).get("body", [])
    return {item["NAME"]: str(item["CODE"]) for item in body}


def fetch_sgg_city_map(session: requests.Session, city_code: str, election_code: str) -> dict[str, str]:
    """{시군구명: 코드}. sggCityCode JSON. 일반시(수원시 등)가 여기 들어있다."""
    r = session.post(
        f"{BASE}/bizcommon/selectbox/selectbox_getSggCityCodeJson.json",
        data={"electionId": ELECTION_ID, "electionCode": election_code, "cityCode": city_code},
        timeout=30,
    )
    body = r.json().get("jsonResult", {}).get("body", [])
    # sggCityCode 는 7자리(예: 4410300). 사진 디렉토리는 4자리 코드 필요 → 앞 4자리 사용
    return {item["NAME"]: str(item["CODE"])[:4] for item in body}


def build_all_maps() -> dict[str, dict[str, str]]:
    session = build_session()
    out: dict[str, dict[str, str]] = {}
    sido_codes = set(SIDO_TO_CODE.values())
    for sido_code in sido_codes:
        print(f"  매핑 다운로드: {sido_code}")
        m6 = fetch_town_map(session, sido_code, "6")   # 일반 townCode (구단위)
        time.sleep(0.12)
        m5 = fetch_town_map(session, sido_code, "5")   # 통합 townCode
        time.sleep(0.12)
        m4 = fetch_sgg_city_map(session, sido_code, "4")  # sggCityCode (시단위, 일반시 포함)
        time.sleep(0.12)
        merged = {**m4, **m6, **m5}
        if merged:
            out[sido_code] = merged
    # 통합특별시(2900)에 전남(4600) 매핑도 머지 (시도의원 5번 통합권역 대응)
    if "2900" in out and "4600" in out:
        out["2900"] = {**out["4600"], **out["2900"]}
    return out


def extract_district(sgg: str, map_dict: dict[str, str]) -> str | None:
    """매핑 사전에서 가장 긴 prefix/suffix 매칭을 찾는다. 없으면 None."""
    if not sgg:
        return None
    sgg = sgg.strip()
    # 긴 이름부터 매칭 (예: '수원시장안구' > '수원시')
    names = sorted(map_dict.keys(), key=len, reverse=True)
    for name in names:
        if name in sgg:
            return name
    return None


def main():
    print("1) 시도별 시군구 매핑 다운로드")
    maps = build_all_maps()
    print(f"   완료 — 시도 {len(maps)}개")
    for code, m in maps.items():
        print(f"   {code}: {len(m)}개 시군구 (sample: {list(m.items())[:3]})")

    print("\n2) 후보별 photo URL 재구성")
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["전체"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    urls: dict[str, str] = {}
    miss_examples = []
    stats = {"sido_level": 0, "town_ok": 0, "town_miss": 0, "no_sgg": 0}

    for r in ws.iter_rows(min_row=2, values_only=True):
        hubo = str(r[idx["huboId"]])
        elec = r[idx["선거종류"]]
        sido = r[idx["시도"]]
        sgg = r[idx["선거구"]] or ""
        if not hubo:
            continue
        sido_code = SIDO_TO_CODE.get(sido)
        if not sido_code:
            continue
        # 시도 단위 선거
        if elec in SIDO_LEVEL_ELECTIONS:
            code = sido_code
            stats["sido_level"] += 1
        else:
            sido_map = maps.get(sido_code, {})
            dist = extract_district(sgg, sido_map)
            if not dist:
                stats["town_miss"] += 1
                if len(miss_examples) < 10:
                    miss_examples.append(f"{sido}/{sgg!r}")
                continue
            code = sido_map[dist]
            stats["town_ok"] += 1
        urls[hubo] = f"{BASE}/photo_20260603/Gsg{code}/Hb{hubo}/gicho/{hubo}.JPG"

    OUT.write_text(json.dumps(urls, ensure_ascii=False), encoding="utf-8")
    print(f"\n완료 — {len(urls)}건 URL 재구성")
    print(f"  시도단위: {stats['sido_level']}, 구시군매칭OK: {stats['town_ok']}, "
          f"구시군매칭실패: {stats['town_miss']}, sgg없음(시도fallback): {stats['no_sgg']}")
    if miss_examples:
        print("\n매칭 실패 샘플:")
        for ex in miss_examples:
            print(f"  - {ex}")
    print(f"\n저장: {OUT}")


if __name__ == "__main__":
    main()
