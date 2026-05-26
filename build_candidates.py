"""
후보자별 단일 HTML 페이지 생성
- 입력: nec_2026_지방선거_후보자.xlsx, .photo_db.json, pledges.json (선택: pledges_text.json)
- 출력: candidates/{huboId}.html (≈ 7,826개), candidates/style.css, candidates/qa.js
- 각 페이지 구성:
    · 헤더(사진 + 이름 + 정당/기호/선거구/선거종류)
    · 기본정보 + 학력/경력 + 재산·병역·납세·전과(xlsx 요약)
    · "NEC 상세보기" 버튼 (재산/병역/납세/전과/학력/공직선거경력 스캔 PDF 보기)
    · 선거공보 / 5대공약 PDF iframe (cdn.nec.go.kr)
    · Gemini Q&A 챗 (사용자가 직접 키 입력 / 또는 빌드 시 키 임베드)
"""

from __future__ import annotations

import argparse
import html
import json
import logging
import re
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("build_candidates")

NEC_DETAIL_URL = "https://info.nec.go.kr/electioninfo/candidate_detail_info.xhtml?electionId=0020260603&huboId={huboId}"

# 파티 색상 (build_dashboard.py와 동일)
PARTY_COLORS = {
    "더불어민주당": "#004EA2", "국민의힘": "#E61E2B", "조국혁신당": "#0073CF",
    "개혁신당": "#FF7920", "진보당": "#D6001C", "정의당": "#FFCC00",
    "기본소득당": "#00B5A8", "새미래민주당": "#1AA75E", "자유와혁신": "#8A2BE2",
    "노동당": "#EB121A", "자유통일당": "#A6324D", "무소속": "#9ca3af",
}


def normalize_party(p: str | None) -> str:
    if not p:
        return "무소속"
    return re.sub(r"\(\d+\)\s*$", "", str(p)).strip() or "무소속"


def party_color(p: str) -> str:
    return PARTY_COLORS.get(p, "#7c3aed")


def load_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["전체"]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r): continue
        rows.append(dict(zip(headers, r)))
    log.info("xlsx loaded: %d rows", len(rows))
    return rows


SHARED_CSS = """
:root {
  --primary: #5645d4; --primary-pressed: #4534b3; --on-primary: #ffffff;
  --brand-navy: #0a1530; --link-blue: #0075de;
  --tint-peach:#ffe8d4; --tint-rose:#fde0ec; --tint-mint:#d9f3e1; --tint-lavender:#e6e0f5; --tint-sky:#dcecfa;
  --canvas:#ffffff; --surface:#f6f5f4; --surface-soft:#fafaf9;
  --hairline:#e5e3df; --hairline-soft:#ede9e4;
  --ink:#1a1a1a; --charcoal:#37352f; --slate:#5d5b54; --steel:#787671; --stone:#a4a097; --muted:#bbb8b1;
  --shadow-card: rgba(15,15,15,0.06) 0px 4px 12px 0px;
  --r-sm:6px; --r-md:8px; --r-lg:12px; --r-xl:16px; --r-full:9999px;
}
*{box-sizing:border-box}
html,body{margin:0;padding:0}
body{
  font-family:'Inter',-apple-system,BlinkMacSystemFont,'Apple SD Gothic Neo','맑은 고딕',sans-serif;
  background:var(--surface); color:var(--ink); font-size:15px; line-height:1.55;
  -webkit-font-smoothing:antialiased;
}
a{color:var(--link-blue); text-decoration:none}
a:hover{text-decoration:underline}
.container{max-width:960px; margin:0 auto; padding:24px 20px 80px}
.topbar{display:flex; align-items:center; gap:12px; padding:14px 0; font-size:13px; color:var(--steel)}
.topbar a{color:var(--steel)}
.topbar .sep{color:var(--muted)}

/* Header card */
.hero-card{
  background:var(--canvas); border:1px solid var(--hairline); border-radius:var(--r-xl);
  padding:24px; display:grid; grid-template-columns:120px 1fr; gap:24px; align-items:start;
  box-shadow:var(--shadow-card);
}
.hero-photo{
  width:120px; height:160px; border-radius:var(--r-md); object-fit:cover; background:var(--surface);
  border:1px solid var(--hairline);
}
.hero-photo.missing{
  display:flex; align-items:center; justify-content:center; color:var(--muted); font-size:32px;
}
.hero-meta{display:flex; flex-direction:column; gap:6px}
.hero-name{font-size:28px; font-weight:700; color:var(--charcoal); letter-spacing:-0.01em}
.hero-name .hanja{color:var(--steel); font-weight:400; font-size:18px; margin-left:8px}
.hero-tags{display:flex; flex-wrap:wrap; gap:6px; margin-top:4px}
.tag{display:inline-flex; align-items:center; gap:4px; padding:3px 10px; border-radius:var(--r-full); font-size:12px; font-weight:600}
.tag.giho{background:var(--surface); color:var(--charcoal); border:1px solid var(--hairline)}
.tag.sg{background:var(--tint-sky); color:#1d4ed8}
.tag.region{background:var(--tint-mint); color:#15803d}
.hero-sub{color:var(--slate); font-size:14px; margin-top:6px}

/* Sections */
.section{
  background:var(--canvas); border:1px solid var(--hairline); border-radius:var(--r-lg);
  padding:20px 22px; margin-top:18px;
}
.section h2{font-size:16px; margin:0 0 14px; color:var(--charcoal); display:flex; align-items:center; gap:8px}
.section h2 .badge{font-size:11px; color:var(--steel); font-weight:500; background:var(--surface); padding:2px 8px; border-radius:var(--r-full)}
.row{display:grid; grid-template-columns:96px 1fr; gap:10px; padding:8px 0; border-bottom:1px solid var(--hairline-soft); font-size:14px}
.row:last-child{border-bottom:none}
.row .k{color:var(--steel); font-weight:500}
.row .v{color:var(--charcoal); white-space:pre-wrap; word-break:break-word}
.row .v.empty{color:var(--muted)}

.nec-link-btn{
  display:inline-flex; align-items:center; gap:6px; padding:8px 14px;
  background:var(--surface); border:1px solid var(--hairline); border-radius:var(--r-md);
  color:var(--charcoal); font-size:13px; font-weight:600; cursor:pointer;
  margin-top:14px;
}
.nec-link-btn:hover{background:var(--surface-soft); text-decoration:none}
.nec-link-btn .ico{color:var(--primary)}

/* PDF embed */
.pdf-wrap{margin-top:10px; border:1px solid var(--hairline); border-radius:var(--r-md); overflow:hidden; background:#525252; position:relative; min-height:480px}
.pdf-wrap iframe{display:block; width:100%; height:720px; border:0; background:#525252}
.pdf-loading{position:absolute; inset:0; display:flex; align-items:center; justify-content:center; color:#fff; font-size:13px; pointer-events:none; transition:opacity 0.2s}
.pdf-loading.hidden{opacity:0}
.pdf-tabs{display:flex; gap:0; border-bottom:1px solid var(--hairline); background:var(--canvas)}
.pdf-tab{padding:10px 16px; cursor:pointer; font-size:13px; font-weight:600; color:var(--steel); border-bottom:2px solid transparent; user-select:none}
.pdf-tab:hover{color:var(--charcoal)}
.pdf-tab.active{color:var(--primary); border-bottom-color:var(--primary)}
.pdf-empty{padding:32px 20px; color:var(--muted); text-align:center; font-size:14px}
.pdf-download-row{display:flex; gap:10px; flex-wrap:wrap; margin-top:14px}
.pdf-download{display:inline-flex; align-items:center; gap:8px; padding:10px 16px; background:var(--canvas); border:1px solid var(--hairline); border-radius:var(--r-md); color:var(--charcoal); font-size:13px; font-weight:600; text-decoration:none; transition:border-color 0.12s ease}
.pdf-download:hover{border-color:var(--primary); color:var(--primary); text-decoration:none}
.pdf-download .ico{font-size:14px; color:var(--primary)}

/* Q&A */
.qa-box{margin-top:14px}
.qa-intro{font-size:13px; color:var(--steel); margin-bottom:10px; line-height:1.6}
.qa-msgs{display:flex; flex-direction:column; gap:8px; margin-bottom:12px; max-height:520px; overflow-y:auto; padding-right:4px}
.qa-msg{padding:10px 14px; border-radius:var(--r-md); font-size:14px; line-height:1.6; white-space:pre-wrap; word-break:break-word}
.qa-msg.user{background:var(--primary); color:var(--on-primary); align-self:flex-end; max-width:80%}
.qa-msg.bot{background:var(--surface); color:var(--charcoal); align-self:flex-start; max-width:90%; border:1px solid var(--hairline)}
.qa-msg.bot.error{background:#fff1f0; border-color:#fecaca; color:#991b1b}
.qa-msg.bot.loading{color:var(--steel); font-style:italic}
.qa-input-row{display:flex; gap:8px}
.qa-input-row input{flex:1; padding:11px 14px; border:1px solid var(--hairline); border-radius:var(--r-md); font-family:inherit; font-size:14px}
.qa-input-row input:focus{outline:none; border-color:var(--primary)}
.qa-send{padding:11px 18px; background:var(--primary); color:var(--on-primary); border:none; border-radius:var(--r-md); font-weight:600; cursor:pointer; font-family:inherit; font-size:14px}
.qa-send:hover:not(:disabled){background:var(--primary-pressed)}
.qa-send:disabled{opacity:0.5; cursor:not-allowed}
.qa-suggestions{display:flex; flex-wrap:wrap; gap:6px; margin-bottom:10px}
.qa-chip{padding:5px 11px; background:var(--surface); border:1px solid var(--hairline); border-radius:var(--r-full); font-size:12px; color:var(--charcoal); cursor:pointer}
.qa-chip:hover{background:var(--surface-soft); border-color:var(--primary)}

@media (max-width: 640px){
  .hero-card{grid-template-columns:80px 1fr; padding:16px; gap:14px}
  .hero-photo{width:80px; height:106px}
  .hero-name{font-size:22px}
  .row{grid-template-columns:80px 1fr}
  .pdf-wrap iframe{height:520px}
}
"""


QA_JS = r"""
// 후보자 페이지 클라이언트 로직
//   1) PDF 미리보기: NEC CDN 은 application/x-pdf 로 보내서 iframe 직접 src 가 안 먹는다.
//      → fetch → Blob(application/pdf) → URL.createObjectURL → iframe.src
//   2) Q&A: POST /api/qa (Vercel 서버리스). 키는 서버의 GEMINI_API_KEY 에서 읽음.
// 데이터:  window.CANDIDATE = {huboId, name, party, ...}
//          window.PLEDGE_TEXT = "..."  (선택: 사전 추출 텍스트)
//          window.PLEDGE_PDFS = [{type, url}, ...]
(function(){
  const $ = sel => document.querySelector(sel);
  const $$ = sel => document.querySelectorAll(sel);

  // ----- PDF preview (blob URL) -----
  const blobCache = new Map();          // url → blob URL
  async function loadPdfIntoFrame(url){
    const frame = $('#pdfFrame');
    const loading = $('#pdfLoading');
    if (!frame) return;
    if (loading){ loading.classList.remove('hidden'); loading.textContent = 'PDF 불러오는 중…'; }
    try{
      if (!blobCache.has(url)){
        const r = await fetch(url);
        if (!r.ok) throw new Error('HTTP '+r.status);
        const buf = await r.arrayBuffer();
        const blob = new Blob([buf], { type: 'application/pdf' });
        blobCache.set(url, URL.createObjectURL(blob));
      }
      frame.src = blobCache.get(url);
      if (loading){
        setTimeout(() => loading.classList.add('hidden'), 200);
      }
    }catch(e){
      if (loading){ loading.textContent = 'PDF 로딩 실패: '+(e.message||e); }
    }
  }

  function setActiveTab(tab){
    $$('.pdf-tab').forEach(t => t.classList.remove('active'));
    tab.classList.add('active');
    const url = tab.dataset.pdf;
    if (url) loadPdfIntoFrame(url);
    // download button updates to currently active PDF
    const dl = $('#pdfDownloadActive');
    if (dl){
      dl.href = url;
      dl.download = (window.CANDIDATE?.name || 'pledge') + '_' + tab.textContent.trim() + '.pdf';
      dl.querySelector('.label').textContent = tab.textContent.trim() + ' PDF 다운로드';
    }
  }

  // ----- Q&A (uses /api/qa) -----
  function addMsg(role, text, extraClass=''){
    const div = document.createElement('div');
    div.className = 'qa-msg ' + role + (extraClass ? ' '+extraClass : '');
    div.textContent = text;
    const msgs = $('#qaMsgs');
    msgs.appendChild(div);
    msgs.scrollTop = msgs.scrollHeight;
    return div;
  }

  async function ask(question){
    addMsg('user', question);
    const loading = addMsg('bot', '공약 자료를 읽고 답변 작성 중…', 'loading');
    $('#qaInput').value = '';
    $('#qaSend').disabled = true;

    try{
      const r = await fetch('/api/qa', {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({
          question,
          candidate: window.CANDIDATE,
          pledgeText: window.PLEDGE_TEXT || '',
          pdfUrls: window.PLEDGE_PDFS || []
        })
      });
      const j = await r.json().catch(() => ({error:'응답 파싱 실패'}));
      loading.remove();
      if (!r.ok){
        addMsg('bot', j.error || ('서버 오류 '+r.status), 'error');
        return;
      }
      addMsg('bot', (j.answer || '').trim() || '(빈 응답)');
    } catch(e){
      loading.remove();
      addMsg('bot', '요청 실패: '+(e.message||e), 'error');
    } finally {
      $('#qaSend').disabled = false;
    }
  }

  window.addEventListener('DOMContentLoaded', () => {
    // PDF: first tab 자동 로드
    const first = $('.pdf-tab.active') || $('.pdf-tab');
    if (first) setActiveTab(first);
    $$('.pdf-tab').forEach(tab => tab.addEventListener('click', () => setActiveTab(tab)));

    // Q&A 폼
    const form = $('#qaForm');
    if (form){
      form.addEventListener('submit', e => {
        e.preventDefault();
        const q = $('#qaInput').value.trim();
        if (q) ask(q);
      });
      $$('.qa-chip').forEach(chip => {
        chip.addEventListener('click', () => ask(chip.textContent));
      });
    }
  });
})();
"""


def fmt_won(thousands: int | str | None) -> str:
    """천원 단위 정수 → '12억 3,456만 7천원' 한국어 표기"""
    if thousands in (None, "", 0, "0"):
        return "0원"
    try:
        n = int(str(thousands).replace(",", ""))
    except ValueError:
        return str(thousands)
    if n == 0: return "0원"
    sign = "-" if n < 0 else ""
    n = abs(n)
    won = n * 1000  # 천원→원
    eok, rem = divmod(won, 100_000_000)
    man, rem = divmod(rem, 10_000)
    parts = []
    if eok: parts.append(f"{eok:,}억")
    if man: parts.append(f"{man:,}만")
    if rem and not eok: parts.append(f"{rem:,}")
    if not parts: parts.append("0")
    return sign + " ".join(parts) + "원"


def render_page(cand: dict, photo: dict | None, pledge: dict | None,
                pledge_text: str | None) -> str:
    h = html.escape
    name = cand.get("성명") or ""
    hanja = cand.get("한자명") or ""
    party = normalize_party(cand.get("정당"))
    party_c = party_color(party)
    sgg = cand.get("선거구") or ""
    sg_name = cand.get("선거종류") or ""
    sido = cand.get("시도") or ""
    giho = cand.get("기호") or ""
    hubo_id = str(cand.get("huboId") or "")

    photo_url = ""
    if photo and photo.get("has_photo"):
        photo_url = f"../photos/{hubo_id}.jpg"

    # 기본 정보 row
    def row(k, v, *, empty="—"):
        v_s = str(v).strip() if v else ""
        if not v_s:
            return f'<div class="row"><div class="k">{h(k)}</div><div class="v empty">{empty}</div></div>'
        return f'<div class="row"><div class="k">{h(k)}</div><div class="v">{h(v_s)}</div></div>'

    basic_rows = [
        row("성별", cand.get("성별")),
        row("생년월일", cand.get("생년월일")),
        row("직업", cand.get("직업")),
        row("주소", cand.get("주소")),
    ]
    edu_rows = [
        row("학력", cand.get("학력"), empty="제출된 학력 정보 없음"),
        row("경력", cand.get("경력"), empty="제출된 경력 정보 없음"),
    ]

    asset = cand.get("재산신고액(천원)")
    asset_disp = fmt_won(asset) if asset is not None else "—"
    tax_paid = cand.get("납세_납부액(천원)")
    tax_5y = cand.get("납세_5년체납(천원)")
    tax_now = cand.get("납세_현체납(천원)")

    finance_rows = [
        row("재산신고액", asset_disp + (f"  (원자료: {asset:,}천원)" if isinstance(asset, (int,float)) and asset else "")),
        row("최근 5년 납부세액", fmt_won(tax_paid) if tax_paid is not None else "—"),
        row("최근 5년 체납액", fmt_won(tax_5y) if tax_5y is not None else "—"),
        row("현재 체납액", fmt_won(tax_now) if tax_now is not None else "—"),
        row("병역", cand.get("병역")),
        row("전과", cand.get("전과"), empty="없음 (또는 미공개)"),
        row("입후보 횟수", cand.get("입후보횟수")),
    ]

    # 공약 PDF 탭
    pdfs = []
    if pledge:
        files = pledge.get("files") or {}
        for typ, url in files.items():
            pdfs.append({"type": typ, "url": url})
    pdf_section_html = ""
    if pdfs:
        tabs_html = "".join(
            f'<div class="pdf-tab{ " active" if i==0 else "" }" data-pdf="{h(p["url"])}">{h(p["type"])}</div>'
            for i, p in enumerate(pdfs)
        )
        first = pdfs[0]
        first_dl = f"{name}_{first['type']}.pdf"
        # active 다운로드 버튼 (탭 따라 갱신) + 모든 PDF 별도 직링크
        all_downloads = "".join(
            f'<a class="pdf-download" href="{h(p["url"])}" download="{h(name)}_{h(p["type"])}.pdf" target="_blank" rel="noopener">'
            f'<span class="ico">↓</span><span>{h(p["type"])} PDF 다운로드</span></a>'
            for p in pdfs
        )
        pdf_section_html = f"""
        <div class="pdf-tabs">{tabs_html}</div>
        <div class="pdf-wrap">
          <iframe id="pdfFrame" title="후보자 공약 PDF"></iframe>
          <div id="pdfLoading" class="pdf-loading">PDF 불러오는 중…</div>
        </div>
        <div class="pdf-download-row">{all_downloads}</div>
        """
    else:
        pdf_section_html = '<div class="pdf-empty">중앙선관위에 제출된 공약 자료가 없습니다.</div>'

    # 질문 추천
    suggestions = [
        "이 후보의 핵심 5대 공약은?",
        "청년·일자리 관련 공약을 정리해줘",
        "주거·부동산 정책은?",
        "재원 조달 방안이 명시돼 있어?",
    ]
    chips_html = "".join(f'<button type="button" class="qa-chip">{h(s)}</button>' for s in suggestions)

    # 임베드 데이터
    candidate_js = json.dumps({
        "huboId": hubo_id, "name": name, "party": party, "giho": giho,
        "subSgName": sg_name, "sgg": sgg, "sido": sido,
    }, ensure_ascii=False)
    pdfs_js = json.dumps(pdfs, ensure_ascii=False)
    pledge_text_js = json.dumps(pledge_text or "", ensure_ascii=False)

    nec_link = NEC_DETAIL_URL.format(huboId=hubo_id)

    title = f"{name} ({party}) · {sg_name} · {sgg or sido}"

    # 사진 영역
    if photo_url:
        photo_html = f'<img class="hero-photo" src="{h(photo_url)}" alt="{h(name)}" onerror="this.classList.add(\'missing\');this.removeAttribute(\'src\');this.textContent=\' \';">'
    else:
        photo_html = f'<div class="hero-photo missing">사진 없음</div>'

    return f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{h(title)} | 2026 지방선거</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="style.css">
</head>
<body>
<div class="container">
  <div class="topbar">
    <a href="../dashboard.html">← 대시보드로</a>
    <span class="sep">·</span>
    <span>{h(sg_name)} {h(sgg or sido)}</span>
  </div>

  <div class="hero-card">
    {photo_html}
    <div class="hero-meta">
      <div class="hero-name">{h(name)}{f'<span class="hanja">{h(hanja)}</span>' if hanja else ''}</div>
      <div class="hero-tags">
        <span class="tag giho" style="border-color:{party_c}55; color:{party_c}; background:{party_c}11">기호 {h(str(giho))}</span>
        <span class="tag" style="background:{party_c}22; color:{party_c}">{h(party)}</span>
        <span class="tag sg">{h(sg_name)}</span>
        <span class="tag region">{h(sgg or sido)}</span>
      </div>
      <div class="hero-sub">huboId · {h(hubo_id)}</div>
    </div>
  </div>

  <div class="section">
    <h2>기본 정보 <span class="badge">중앙선관위 제출 자료 기준</span></h2>
    {''.join(basic_rows)}
  </div>

  <div class="section">
    <h2>학력 · 경력</h2>
    {''.join(edu_rows)}
  </div>

  <div class="section">
    <h2>재산 · 병역 · 납세 · 전과 <span class="badge">요약</span></h2>
    {''.join(finance_rows)}
    <a class="nec-link-btn" href="{h(nec_link)}" target="_blank" rel="noopener">
      <span class="ico">→</span> 중앙선관위에서 상세 스캔 자료 보기
    </a>
  </div>

  <div class="section">
    <h2>공약 자료 <span class="badge">NEC 정책·공약마당</span></h2>
    {pdf_section_html}
  </div>

  <div class="section">
    <h2>AI 공약 Q&amp;A <span class="badge">Gemini</span></h2>
    <div class="qa-box">
      <div class="qa-intro">
        후보자가 NEC에 제출한 공약 자료를 바탕으로 답변합니다. 사실관계는 원문(위 PDF)을 직접 확인해 주세요.
      </div>
      <div class="qa-suggestions">{chips_html}</div>
      <div class="qa-msgs" id="qaMsgs"></div>
      <form id="qaForm" class="qa-input-row">
        <input id="qaInput" type="text" placeholder="이 후보 공약 중 궁금한 점을 물어보세요…" autocomplete="off">
        <button id="qaSend" class="qa-send" type="submit">질문</button>
      </form>
    </div>
  </div>
</div>

<script>
window.CANDIDATE = {candidate_js};
window.PLEDGE_PDFS = {pdfs_js};
window.PLEDGE_TEXT = {pledge_text_js};
</script>
<script src="qa.js"></script>
</body>
</html>
"""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="nec_2026_지방선거_후보자.xlsx")
    ap.add_argument("--photo-db", default=".photo_db.json")
    ap.add_argument("--pledges", default="pledges.json")
    ap.add_argument("--pledges-text", default="pledges_text.json")
    ap.add_argument("--out-dir", default="candidates")
    ap.add_argument("--limit", type=int, default=0, help="테스트용: 처음 N명만 생성")
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    (out_dir / "style.css").write_text(SHARED_CSS.strip(), encoding="utf-8")
    (out_dir / "qa.js").write_text(QA_JS.strip(), encoding="utf-8")
    log.info("wrote shared style.css & qa.js")

    photo_db = {}
    p = Path(args.photo_db)
    if p.exists():
        photo_db = json.loads(p.read_text(encoding="utf-8"))
        log.info("photo_db: %d entries", len(photo_db))

    pledges = {}
    p = Path(args.pledges)
    if p.exists():
        pledges = json.loads(p.read_text(encoding="utf-8"))
        log.info("pledges: %d entries", len(pledges))
    else:
        log.warning("pledges.json not found; pages will show '공약 없음'")

    pledges_text = {}
    p = Path(args.pledges_text)
    if p.exists():
        pledges_text = json.loads(p.read_text(encoding="utf-8"))
        log.info("pledges_text: %d entries", len(pledges_text))

    candidates = load_xlsx(Path(args.xlsx))
    if args.limit:
        candidates = candidates[: args.limit]

    n = 0
    for cand in candidates:
        hubo_id = str(cand.get("huboId") or "").strip()
        if not hubo_id:
            continue
        html_str = render_page(
            cand=cand,
            photo=photo_db.get(hubo_id),
            pledge=pledges.get(hubo_id),
            pledge_text=pledges_text.get(hubo_id),
        )
        (out_dir / f"{hubo_id}.html").write_text(html_str, encoding="utf-8")
        n += 1
        if n % 1000 == 0:
            log.info("  %d pages written", n)
    log.info("done: %d candidate pages → %s/", n, out_dir)


if __name__ == "__main__":
    main()
