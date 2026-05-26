"""
중앙선거관리위원회 선거통계시스템(info.nec.go.kr) 후보자 크롤러
- 대상 선거: 제9회 전국동시지방선거 (2026-06-03), electionId=0020260603
- 출력: nec_2026_지방선거_후보자.xlsx (선거종류별 시트)
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

BASE = "https://info.nec.go.kr"
ELECTION_ID = "0020260603"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
REQUEST_DELAY = 0.15  # seconds between HTTP calls
TIMEOUT = 30

# 2026 선거에서 시도지사/시도의원/광역비례/교육감(3,5,8,11)은
# "전남광주통합특별시"(cityCode=2900) 같은 통합 광역으로 처리되므로,
# 시도 목록은 하드코딩 대신 API에서 선거코드별로 받아온다.
# 아래는 fallback (네트워크 실패 등) 용 17개 시도.
SIDO_FALLBACK: dict[str, str] = {
    "1100": "서울특별시", "2600": "부산광역시", "2700": "대구광역시",
    "2800": "인천광역시", "2900": "광주광역시", "3000": "대전광역시",
    "3100": "울산광역시", "5100": "세종특별자치시", "4100": "경기도",
    "5200": "강원특별자치도", "4300": "충청북도", "4400": "충청남도",
    "5300": "전북특별자치도", "4600": "전라남도", "4700": "경상북도",
    "4800": "경상남도", "4900": "제주특별자치도",
}

# (electionCode, sheet/election name, drill-down strategy, city list flavor)
# - single       : cityCode 만 사용
# - sgg          : cityCode + sggCityCode (구시군 단위 조회)
# - drill_intg   : cityCode + townCode(통합ver) + sggTownCode (시도의원 5번)
# - drill_normal : cityCode + townCode + sggTownCode (구시군의원 6번)
# city flavor:
# - "di"  : selectbox_cityCodeDIBySgJson.json (통합권역 적용; 3,5,8,11)
# - "reg" : selectbox_cityCodeBySgJson.json   (일반; 2,4,6,9)
ELECTIONS: list[tuple[str, str, str, str]] = [
    ("3",  "시도지사선거",          "single",       "di"),
    ("4",  "구시군의장선거",        "sgg",          "reg"),
    ("5",  "시도의회의원선거",      "drill_intg",   "di"),
    ("6",  "구시군의회의원선거",    "drill_normal", "reg"),
    ("8",  "광역의원비례대표선거",  "single",       "di"),
    ("9",  "기초의원비례대표선거",  "sgg",          "reg"),
    ("11", "교육감선거",            "single",       "di"),
    ("2",  "국회의원선거(보궐)",    "sgg",          "reg"),
]

OUTPUT_HEADERS = [
    "선거종류", "시도", "선거구", "기호", "정당", "성명", "한자명",
    "성별", "생년월일", "주소", "직업", "학력", "경력",
    "재산신고액(천원)", "병역", "납세_납부액(천원)",
    "납세_5년체납(천원)", "납세_현체납(천원)", "전과", "입후보횟수",
    "사진URL", "huboId",
]

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
    level=logging.INFO,
    stream=sys.stdout,
)
log = logging.getLogger("nec")


@dataclass
class Candidate:
    election_name: str
    sido: str
    sgg: str = ""
    gihoNo: str = ""
    party: str = ""
    name: str = ""
    hanja: str = ""
    gender: str = ""
    birth: str = ""
    address: str = ""
    job: str = ""
    education: str = ""
    career: str = ""
    assets: str = ""
    military: str = ""
    tax_paid: str = ""
    tax_5y_delinq: str = ""
    tax_curr_delinq: str = ""
    criminal: str = ""
    run_count: str = ""
    photo_url: str = ""
    hubo_id: str = ""

    def as_row(self) -> list[str]:
        return [
            self.election_name, self.sido, self.sgg, self.gihoNo, self.party,
            self.name, self.hanja, self.gender, self.birth, self.address,
            self.job, self.education, self.career, self.assets, self.military,
            self.tax_paid, self.tax_5y_delinq, self.tax_curr_delinq,
            self.criminal, self.run_count, self.photo_url, self.hubo_id,
        ]


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Referer": f"{BASE}/main/showDocument.xhtml?electionId={ELECTION_ID}&topMenuId=CP&secondMenuId=CPRI03",
        "X-Requested-With": "XMLHttpRequest",
    })
    # bootstrap session cookies (JSESSIONID, WMONID)
    s.get(
        f"{BASE}/main/showDocument.xhtml",
        params={"electionId": ELECTION_ID, "topMenuId": "CP", "secondMenuId": "CPRI03"},
        timeout=TIMEOUT,
    )
    return s


def _post_json(s: requests.Session, path: str, data: dict) -> list[dict]:
    time.sleep(REQUEST_DELAY)
    r = s.post(f"{BASE}{path}", data=data, timeout=TIMEOUT)
    r.raise_for_status()
    try:
        return r.json().get("jsonResult", {}).get("body", []) or []
    except ValueError:
        log.warning("Non-JSON response from %s: %r", path, r.text[:200])
        return []


def get_cities(s, elec_code: str, flavor: str) -> dict[str, str]:
    """선거코드별 시도 목록 (코드→이름). flavor='di' 통합권역, 'reg' 일반."""
    endpoint = ("selectbox_cityCodeDIBySgJson.json" if flavor == "di"
                else "selectbox_cityCodeBySgJson.json")
    try:
        body = _post_json(s, f"/bizcommon/selectbox/{endpoint}",
                          {"electionId": ELECTION_ID, "electionCode": elec_code})
        result = {str(it["CODE"]): it["NAME"] for it in body}
        if result:
            return result
    except requests.HTTPError as e:
        log.warning("cityCode list fetch failed (%s): %s — fallback 사용", flavor, e)
    return dict(SIDO_FALLBACK)


def get_sgg_cities(s, elec_code, city_code) -> list[dict]:
    return _post_json(s, "/bizcommon/selectbox/selectbox_getSggCityCodeJson.json",
                      {"electionId": ELECTION_ID, "electionCode": elec_code, "cityCode": city_code})


def get_towns_intg(s, elec_code, city_code) -> list[dict]:
    return _post_json(s, "/bizcommon/selectbox/selectbox_townCodeByCityIntgSgJson.json",
                      {"electionId": ELECTION_ID, "electionCode": elec_code, "cityCode": city_code})


def get_towns(s, elec_code, city_code) -> list[dict]:
    return _post_json(s, "/bizcommon/selectbox/selectbox_townCodeBySgJson.json",
                      {"electionId": ELECTION_ID, "electionCode": elec_code, "cityCode": city_code})


def get_sgg_towns(s, elec_code, town_code) -> list[dict]:
    return _post_json(s, "/bizcommon/selectbox/selectbox_getSggTownCodeJson.json",
                      {"electionId": ELECTION_ID, "electionCode": elec_code, "townCode": town_code})


def fetch_list_html(s: requests.Session, elec_code: str, city_code: str,
                    extra: dict | None = None) -> str:
    body = {
        "electionId": ELECTION_ID,
        "requestURI": f"/electioninfo/{ELECTION_ID}/cp/cpri03.jsp",
        "topMenuId": "CP", "secondMenuId": "CPRI03", "menuId": "CPRI03",
        "statementId": f"CPRI03_#{elec_code}",
        "electionCode": elec_code,
        "cityCode": city_code,
    }
    if extra:
        body.update(extra)
    time.sleep(REQUEST_DELAY)
    r = s.post(f"{BASE}/electioninfo/electionInfo_report.xhtml",
               data=body, timeout=TIMEOUT)
    r.raise_for_status()
    r.encoding = "utf-8"
    return r.text


# ---------- parsing ---------- #

_HANJA_RE = re.compile(r"[一-鿿]")
_POPUP_RE = re.compile(r"popupHBJ\('([^']+)','(\d+)'\)")
_BIRTH_AGE_RE = re.compile(r"(\d{4}\.\d{2}\.\d{2})")  # birth date


def _cell_text(el) -> str:
    """Get visible text from a cell; <br> becomes newline."""
    if el is None:
        return ""
    for br in el.find_all("br"):
        br.replace_with("\n")
    return el.get_text(strip=True).replace("\xa0", " ")


def _split_name_hanja(text: str) -> tuple[str, str]:
    """'홍길동(洪吉童)' → ('홍길동', '洪吉童')"""
    m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", text)
    if m and _HANJA_RE.search(m.group(2)):
        return m.group(1).strip(), m.group(2).strip()
    return text, ""


def _photo_url(city_code: str, hubo_id: str) -> str:
    return (f"{BASE}/photo_20260603/Gsg{city_code}/"
            f"Hb{hubo_id}/gicho/{hubo_id}.JPG")


def parse_candidate_table(html: str, election_name: str, sido_name: str,
                          city_code: str) -> list[Candidate]:
    """HTML 응답에서 후보자 행을 뽑아 Candidate 리스트로 반환."""
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", id="table01")
    if table is None:
        return []

    # 1) thead 의 컬럼 헤더 추출 (라벨 매핑용)
    headers: list[str] = []
    thead = table.find("thead")
    if thead:
        # 마지막 헤더 row 만 사용 (rowspan/colspan 때문에 여러 줄)
        rows = thead.find_all("tr")
        if rows:
            # 첫 줄과 두번째 줄을 합쳐서 컬럼 순서를 만든다
            first = rows[0].find_all("th")
            second = rows[1].find_all("th") if len(rows) > 1 else []
            second_iter = iter(second)
            for th in first:
                if th.has_attr("colspan") and int(th.get("colspan", "1")) > 1:
                    span = int(th["colspan"])
                    base = _cell_text(th)
                    for _ in range(span):
                        sub = next(second_iter, None)
                        sub_txt = _cell_text(sub) if sub is not None else ""
                        headers.append(f"{base}_{sub_txt}" if sub_txt else base)
                else:
                    headers.append(_cell_text(th))

    # 2) tbody 의 row 들에서 popupHBJ 링크가 있는 행만
    tbody = table.find("tbody") or table
    out: list[Candidate] = []
    for tr in tbody.find_all("tr"):
        a = tr.find("a", href=lambda h: h and "popupHBJ(" in h)
        if a is None:
            continue
        m = _POPUP_RE.search(a["href"])
        hubo_id = m.group(2) if m else ""

        tds = tr.find_all("td")
        if not tds:
            continue
        values = [_cell_text(td) for td in tds]
        col = dict(zip(headers, values))

        name_text = _cell_text(a)
        name, hanja = _split_name_hanja(name_text)

        # 컬럼 라벨에서 안전하게 가져오기
        def g(*keys: str) -> str:
            for k in keys:
                for hk, v in col.items():
                    if k in hk:
                        return v
            return ""

        cand = Candidate(
            election_name=election_name,
            sido=sido_name,
            sgg=g("선거구"),
            gihoNo=g("기호"),
            party=g("정당"),
            name=name,
            hanja=hanja,
            gender=g("성별"),
            birth=g("생년월일", "연령"),
            address=g("주소"),
            job=g("직업"),
            education=g("학력"),
            career=g("경력"),
            assets=g("재산"),
            military=g("병역"),
            tax_paid=g("납부액"),
            tax_5y_delinq=g("5년", "5년간"),
            tax_curr_delinq=g("현체"),
            criminal=g("전과"),
            run_count=g("입후보"),
            photo_url=_photo_url(city_code, hubo_id),
            hubo_id=hubo_id,
        )
        # sgg 컬럼이 "선거구명"으로 시도값(서울특별시 등)이 들어있을 수 있으므로 정리
        if cand.sgg == sido_name:
            cand.sgg = ""
        out.append(cand)
    return out


# ---------- crawl strategies ---------- #

def crawl_election(s: requests.Session, elec_code: str, elec_name: str,
                   strategy: str, city_flavor: str) -> list[Candidate]:
    log.info("=== [%s] %s (strategy=%s, cities=%s) ===",
             elec_code, elec_name, strategy, city_flavor)
    cities = get_cities(s, elec_code, city_flavor)
    log.info("  → 시도 %d개: %s", len(cities), ", ".join(cities.values()))
    out: list[Candidate] = []
    for city_code, city_name in cities.items():
        if strategy == "single":
            html = fetch_list_html(s, elec_code, city_code)
            rows = parse_candidate_table(html, elec_name, city_name, city_code)
            log.info("  [%s] %s → %d명", elec_code, city_name, len(rows))
            out.extend(rows)

        elif strategy == "sgg":
            try:
                sgg_list = get_sgg_cities(s, elec_code, city_code)
            except requests.HTTPError as e:
                log.warning("  sgg city fetch failed %s/%s: %s", elec_code, city_code, e)
                sgg_list = []
            for sgg in sgg_list:
                html = fetch_list_html(s, elec_code, city_code,
                                       {"sggCityCode": sgg["CODE"]})
                rows = parse_candidate_table(html, elec_name, city_name, city_code)
                for r in rows:
                    if not r.sgg:
                        r.sgg = sgg["NAME"]
                out.extend(rows)
            log.info("  [%s] %s 선거구 %d → 누적 %d명",
                     elec_code, city_name, len(sgg_list), len(out))

        elif strategy in ("drill_intg", "drill_normal"):
            town_fetch = get_towns_intg if strategy == "drill_intg" else get_towns
            try:
                towns = town_fetch(s, elec_code, city_code)
            except requests.HTTPError as e:
                log.warning("  towns fetch failed %s/%s: %s", elec_code, city_code, e)
                towns = []
            sgg_count = 0
            cnt_before = len(out)
            for town in towns:
                try:
                    sgg_towns = get_sgg_towns(s, elec_code, town["CODE"])
                except requests.HTTPError as e:
                    log.warning("  sggTowns fetch failed %s/%s: %s",
                                elec_code, town["CODE"], e)
                    sgg_towns = []
                for sgg in sgg_towns:
                    html = fetch_list_html(s, elec_code, city_code, {
                        "townCode": town["CODE"],
                        "sggTownCode": sgg["CODE"],
                    })
                    rows = parse_candidate_table(html, elec_name, city_name, city_code)
                    for r in rows:
                        if not r.sgg:
                            r.sgg = sgg["NAME"]
                    out.extend(rows)
                    sgg_count += 1
            log.info("  [%s] %s 구시군 %d, 선거구 %d → %d명",
                     elec_code, city_name, len(towns), sgg_count,
                     len(out) - cnt_before)
        else:
            log.warning("Unknown strategy: %s", strategy)
    log.info("=== [%s] %s 합계 %d명 ===", elec_code, elec_name, len(out))
    return out


# ---------- excel ---------- #

def write_excel(by_election: dict[str, list[Candidate]], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 전체 합본 시트 먼저
    all_rows = [c for rows in by_election.values() for c in rows]
    _write_sheet(wb, "전체", all_rows)

    for elec_name, rows in by_election.items():
        if rows:
            _write_sheet(wb, elec_name[:31], rows)  # 시트명 31자 제한

    wb.save(path)
    log.info("Excel 저장: %s (총 %d명)", path, len(all_rows))


def _write_sheet(wb: Workbook, name: str, rows: list[Candidate]) -> None:
    ws = wb.create_sheet(name)
    ws.append(OUTPUT_HEADERS)
    for c in rows:
        ws.append(c.as_row())
    # 가독성: 첫 행 굵게, 너비 자동(개략)
    ws.freeze_panes = "A2"
    widths = [10, 12, 18, 6, 16, 10, 10, 6, 14, 30, 14, 36, 36,
              12, 18, 12, 14, 12, 8, 8, 40, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- main ---------- #

def main() -> int:
    ap = argparse.ArgumentParser(description="NEC 2026 지방선거 후보자 크롤러")
    ap.add_argument("--out", type=Path,
                    default=Path(__file__).with_name("nec_2026_지방선거_후보자.xlsx"),
                    help="결과 엑셀 파일 경로")
    ap.add_argument("--only", action="append", default=None,
                    help="특정 선거코드만 (e.g. --only 3 --only 11)")
    args = ap.parse_args()

    s = make_session()
    by_election: dict[str, list[Candidate]] = {}
    for elec_code, elec_name, strategy, city_flavor in ELECTIONS:
        if args.only and elec_code not in args.only:
            continue
        try:
            rows = crawl_election(s, elec_code, elec_name, strategy, city_flavor)
        except KeyboardInterrupt:
            log.warning("중단됨 — 지금까지 수집한 데이터로 저장합니다")
            break
        except Exception:
            log.exception("선거 [%s] %s 처리 중 오류", elec_code, elec_name)
            rows = []
        by_election[elec_name] = rows

    write_excel(by_election, args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
