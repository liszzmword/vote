"""사진 URL 검증 — GET stream으로 헤더만 받아 Content-Type 으로 판정.
결과를 .photo_cache.json 에 저장.
"""
from __future__ import annotations

import json
import time
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).parent
XLSX = ROOT / "nec_2026_지방선거_후보자.xlsx"
CACHE = ROOT / ".photo_cache.json"

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0 Safari/537.36"


def has_real_photo(session: requests.Session, url: str) -> bool:
    """사진 있으면 True, 없으면(=NEC HTML fallback) False."""
    try:
        r = session.get(url, stream=True, timeout=10,
                        headers={"User-Agent": UA, "Accept": "image/*"})
        ct = (r.headers.get("Content-Type") or "").lower()
        cl = int(r.headers.get("Content-Length") or 0)
        r.close()
        if "image" in ct:
            return True
        if "html" in ct:
            return False
        # CT unset → CL 로 보조 판정 (HTML 응답은 3335 고정, 이미지는 더 큼)
        return cl > 6000
    except Exception:
        return None  # unknown — 일단 시도하도록 둠


def load_targets() -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["전체"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        h = str(r[idx["huboId"]])
        url = r[idx["사진URL"]]
        if url:
            out.append((h, url))
    return out


def main():
    cache = {}
    if CACHE.exists():
        cache = json.loads(CACHE.read_text(encoding="utf-8"))
        print(f"기존 캐시 {len(cache)}건 로드")

    targets = load_targets()
    todo = [(h, u) for h, u in targets if h not in cache]
    print(f"전체 {len(targets)}건 / 검증 필요 {len(todo)}건")
    if not todo:
        ok = sum(1 for v in cache.values() if v)
        print(f"OK={ok} NO={len(cache)-ok}")
        return

    done = 0
    last_report = time.time()
    session = requests.Session()

    def work(item):
        h, url = item
        result = has_real_photo(session, url)
        return h, result

    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = [ex.submit(work, t) for t in todo]
        for fut in as_completed(futs):
            try:
                h, ok = fut.result()
                cache[h] = ok
            except Exception:
                pass
            done += 1
            if time.time() - last_report > 5:
                last_report = time.time()
                ok_n = sum(1 for v in cache.values() if v is True)
                no_n = sum(1 for v in cache.values() if v is False)
                un_n = sum(1 for v in cache.values() if v is None)
                print(f"  진행 {done}/{len(todo)} (캐시 OK={ok_n} NO={no_n} UNKNOWN={un_n})")
                CACHE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")

    CACHE.write_text(json.dumps(cache, ensure_ascii=False), encoding="utf-8")
    ok_n = sum(1 for v in cache.values() if v is True)
    no_n = sum(1 for v in cache.values() if v is False)
    un_n = sum(1 for v in cache.values() if v is None)
    print(f"\n완료 — 사진 있음 {ok_n} / 없음 {no_n} / 미확인 {un_n}")
    print(f"캐시 저장: {CACHE} ({CACHE.stat().st_size//1024} KB)")


if __name__ == "__main__":
    main()
