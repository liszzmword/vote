"""사진 DB 빌더.

각 후보의 NEC 상세페이지에서 정확한 photo URL을 추출하고
thumbnail을 photos/{huboId}.jpg 로 다운로드. 결과는 .photo_db.json 에 누적.

- 재실행 안전: 캐시 항목은 skip
- ThreadPoolExecutor(6), request 사이 0.1s sleep
- 5초마다 자동 저장 + 진행률
"""

from __future__ import annotations

import json
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).parent
XLSX = ROOT / "nec_2026_지방선거_후보자.xlsx"
PHOTOS_DIR = ROOT / "photos"
DB_FILE = ROOT / ".photo_db.json"

BASE = "https://info.nec.go.kr"
ELECTION_ID = "0020260603"
UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)

PHOTO_RE = re.compile(r"/photo_20260603/[^\"'?]+\.JPG", re.IGNORECASE)


def thumb_from_full(url: str) -> str:
    return re.sub(r"/gicho/(\d+)\.JPG$", r"/gicho/thumbnail.\1.JPG", url, flags=re.IGNORECASE)


def fetch_detail_photo_url(session: requests.Session, hubo_id: str) -> str | None:
    """NEC 상세페이지에서 정확한 풀사이즈 photo URL을 추출. 없으면 None."""
    url = f"{BASE}/electioninfo/candidate_detail_info.xhtml?electionId={ELECTION_ID}&huboId={hubo_id}"
    try:
        r = session.get(url, timeout=15, headers={"User-Agent": UA})
        if r.status_code != 200:
            return None
        m = PHOTO_RE.search(r.text)
        return BASE + m.group(0) if m else None
    except Exception:
        return None


def download_thumb(session: requests.Session, thumb_url: str, dest: Path) -> bool:
    """썸네일 다운로드. image/* 이고 크기 ≥ 5KB 이면 저장하고 True."""
    try:
        r = session.get(thumb_url, timeout=15, stream=True,
                        headers={"User-Agent": UA, "Accept": "image/*"})
        if r.status_code != 200:
            return False
        ct = (r.headers.get("Content-Type") or "").lower()
        if "image" not in ct:
            r.close()
            return False
        content = r.content
        if len(content) < 5000:
            return False
        dest.write_bytes(content)
        return True
    except Exception:
        return False


def process_one(session: requests.Session, hubo_id: str) -> dict:
    """후보 1명 처리 → 캐시 항목 dict 반환."""
    full_url = fetch_detail_photo_url(session, hubo_id)
    if not full_url:
        return {"url": None, "thumb_url": None, "local": None, "has_photo": False}
    thumb_url = thumb_from_full(full_url)
    dest = PHOTOS_DIR / f"{hubo_id}.jpg"
    ok = download_thumb(session, thumb_url, dest)
    return {
        "url": full_url,
        "thumb_url": thumb_url,
        "local": f"photos/{hubo_id}.jpg" if ok else None,
        "has_photo": ok,
    }


def load_huboids() -> list[str]:
    if not XLSX.exists():
        sys.exit(f"xlsx 없음: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["전체"]
    hdr = [c.value for c in ws[1]]
    i = hdr.index("huboId")
    return [str(r[i]) for r in ws.iter_rows(min_row=2, values_only=True) if r[i]]


def main():
    PHOTOS_DIR.mkdir(exist_ok=True)
    db: dict[str, dict] = {}
    if DB_FILE.exists():
        try:
            db = json.loads(DB_FILE.read_text(encoding="utf-8"))
            print(f"기존 캐시 {len(db)}건 로드")
        except Exception as e:
            print(f"캐시 로드 실패 ({e}) — 비우고 시작")
            db = {}

    targets = load_huboids()
    # 캐시에 None(미확인) 이거나 없는 항목만 재처리
    todo = [h for h in targets
            if h not in db or db[h].get("has_photo") is None]
    print(f"전체 {len(targets)}건 / 처리 필요 {len(todo)}건")
    if not todo:
        _summary(db, len(targets))
        return

    session = requests.Session()
    last_save = time.time()
    done = 0

    with ThreadPoolExecutor(max_workers=6) as ex:
        futs = {ex.submit(process_one, session, h): h for h in todo}
        for fut in as_completed(futs):
            h = futs[fut]
            try:
                db[h] = fut.result()
            except Exception as e:
                db[h] = {"url": None, "thumb_url": None, "local": None, "has_photo": None}
            done += 1
            # 작은 sleep 으로 부하 완화 (각 worker thread 가 받음)
            time.sleep(0.05)
            if time.time() - last_save > 5:
                last_save = time.time()
                _save(db)
                ok = sum(1 for v in db.values() if v.get("has_photo") is True)
                no = sum(1 for v in db.values() if v.get("has_photo") is False)
                un = sum(1 for v in db.values() if v.get("has_photo") is None)
                print(f"  진행 {done}/{len(todo)} (캐시 OK={ok} NO={no} UNKNOWN={un})")

    _save(db)
    _summary(db, len(targets))


def _save(db: dict) -> None:
    DB_FILE.write_text(json.dumps(db, ensure_ascii=False), encoding="utf-8")


def _summary(db: dict, total: int) -> None:
    ok = sum(1 for v in db.values() if v.get("has_photo") is True)
    no = sum(1 for v in db.values() if v.get("has_photo") is False)
    un = sum(1 for v in db.values() if v.get("has_photo") is None)
    print(f"\n완료 — 사진 있음 {ok} / 없음 {no} / 미확인 {un} / 합계 {len(db)} (전체 {total})")
    if PHOTOS_DIR.exists():
        files = list(PHOTOS_DIR.glob("*.jpg"))
        total_bytes = sum(p.stat().st_size for p in files)
        print(f"photos/: {len(files)}장, {total_bytes // (1024*1024)} MB")


if __name__ == "__main__":
    main()
