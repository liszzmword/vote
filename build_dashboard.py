"""
2026 지방선거 후보자 분석 대시보드 — Notion 스타일 디자인 시스템 적용.
크롤링된 xlsx 를 단일 dashboard.html 로 변환 (데이터 임베드, 오프라인 동작).
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
XLSX = ROOT / "nec_2026_지방선거_후보자.xlsx"
OUT = ROOT / "dashboard.html"
PHOTO_DB = ROOT / ".photo_db.json"          # huboId -> {url, thumb_url, local, has_photo} (1순위)
PHOTO_URLS = ROOT / ".photo_urls.json"      # huboId -> 재구성된 사진 URL (fallback)
PHOTO_CACHE = ROOT / ".photo_cache.json"    # 더이상 사용 안 함 (호환용)


# ---------- field parsing ---------- #

def parse_assets(s) -> int | None:
    if s is None:
        return None
    txt = str(s).strip()
    if not txt:
        return None
    cleaned = re.sub(r"[^\d\-]", "", txt)
    if not cleaned or cleaned == "-":
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def parse_age(s) -> int | None:
    if not s:
        return None
    m = re.search(r"\((\d+)\s*세\)", str(s))
    return int(m.group(1)) if m else None


def parse_birth_date(s):
    """'1968.08.12(57세)' → '1968.08.12'"""
    if not s:
        return ""
    m = re.match(r"(\d{4}\.\d{2}\.\d{2})", str(s).strip())
    return m.group(1) if m else str(s).strip()


def parse_criminal(s) -> int:
    if not s:
        return 0
    txt = str(s)
    if "없음" in txt:
        return 0
    m = re.search(r"(\d+)\s*건", txt)
    return int(m.group(1)) if m else 0


def normalize_party(p) -> str:
    if p is None or str(p).strip() == "":
        return "무소속"
    return re.sub(r"\(\d+\)\s*$", "", str(p)).strip() or "무소속"


def categorize_military(s) -> str:
    if not s:
        return "해당없음"
    txt = str(s)
    if "마친" in txt:
        return "복무필"
    if "마치지" in txt or "미필" in txt:
        return "미필"
    if "면제" in txt:
        return "면제"
    if "비대상" in txt or "해당없음" in txt:
        return "비대상"
    if "복무중" in txt or "복무 중" in txt:
        return "복무중"
    if "병적기록이 없" in txt:
        return "병적없음"
    return "기타"


def categorize_job(s) -> str:
    if not s:
        return "(미상)"
    txt = str(s).strip()
    rules = [
        ("국회의원", ["국회의원"]),
        ("지방자치단체장", ["시장", "구청장", "군수", "도지사", "지사"]),
        ("지방의원", ["의원", "의장"]),
        ("정당인", ["정당인", "당원", "최고위원", "대변인"]),
        ("교육자", ["교수", "교사", "강사", "교장", "총장", "학장"]),
        ("법조인", ["변호사", "검사", "판사", "법무사", "노무사"]),
        ("의료인", ["의사", "한의사", "치과", "약사", "간호"]),
        ("회사원/경영", ["회장", "대표", "사장", "이사", "사업", "임원", "CEO", "회사원"]),
        ("공무원", ["공무원"]),
        ("농어업", ["농업", "어업", "축산", "임업"]),
        ("자영업", ["자영", "상업", "운수"]),
        ("종교인", ["목사", "스님", "신부", "전도사"]),
        ("언론/문화/예술", ["기자", "PD", "방송", "예술", "작가", "감독"]),
        ("무직/기타", ["무직", "주부"]),
    ]
    for label, keys in rules:
        if any(k in txt for k in keys):
            return label
    return "기타"


# ---------- load ---------- #

def load_records() -> list[dict]:
    if not XLSX.exists():
        sys.exit(f"입력 파일 없음: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["전체"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}

    # 1) photo_db (로컬 파일) — 최우선
    photo_db: dict[str, dict] = {}
    if PHOTO_DB.exists():
        try:
            photo_db = json.loads(PHOTO_DB.read_text(encoding="utf-8"))
            ok = sum(1 for v in photo_db.values() if v.get("has_photo") is True)
            no = sum(1 for v in photo_db.values() if v.get("has_photo") is False)
            print(f"photo_db 적용: 있음 {ok} / 없음 {no} (로컬 파일 우선)")
        except Exception as e:
            print(f"photo_db 로드 실패 ({e})")

    # 2) 재구성 URL 캐시 — fallback
    url_cache: dict[str, str] = {}
    if PHOTO_URLS.exists():
        try:
            url_cache = json.loads(PHOTO_URLS.read_text(encoding="utf-8"))
            print(f"재구성 URL 캐시 (fallback): {len(url_cache)}건")
        except Exception as e:
            print(f"URL 캐시 로드 실패 ({e})")

    # 3) 공약 PDF URL 매핑 — Q&A에 활용
    pledges_db: dict[str, dict] = {}
    pledges_path = ROOT / "pledges.json"
    if pledges_path.exists():
        try:
            full = json.loads(pledges_path.read_text(encoding="utf-8"))
            # compact: huboId → [{type, url}, ...]
            for hid, p in full.items():
                files = p.get("files") or {}
                if files:
                    pledges_db[hid] = [{"type": t, "url": u} for t, u in files.items()]
            print(f"공약 PDF 매핑: {len(pledges_db)}건")
        except Exception as e:
            print(f"pledges.json 로드 실패 ({e})")

    def col(row, name, default=""):
        v = row[idx[name]] if name in idx else None
        return v if v is not None else default

    out: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        hubo_id = str(col(r, "huboId"))
        # 우선순위: photo_db(로컬 파일) > photo_urls(재구성) > xlsx 원본
        db_entry = photo_db.get(hubo_id)
        if db_entry is not None:
            if db_entry.get("has_photo"):
                ph = db_entry.get("local") or ""       # "photos/100157144.jpg"
            elif db_entry.get("has_photo") is False:
                ph = ""                                # 사진 없음 (placeholder)
            else:  # has_photo == None → 미확인 → fallback
                ph = url_cache.get(hubo_id) or col(r, "사진URL")
        else:
            ph = url_cache.get(hubo_id) or col(r, "사진URL")
        out.append({
            "e": col(r, "선거종류"),
            "s": col(r, "시도"),
            "g": col(r, "선거구"),
            "gh": col(r, "기호"),
            "p": normalize_party(col(r, "정당", None)),
            "n": col(r, "성명"),
            "hj": col(r, "한자명"),
            "gd": col(r, "성별"),
            "a": parse_age(col(r, "생년월일")),
            "bd": parse_birth_date(col(r, "생년월일")),
            "addr": col(r, "주소"),
            "j": categorize_job(col(r, "직업")),
            "jr": col(r, "직업"),
            "edu": col(r, "학력"),
            "car": col(r, "경력"),
            "as": parse_assets(col(r, "재산신고액(천원)")),
            "m": categorize_military(col(r, "병역")),
            "mr": col(r, "병역"),
            "tp": parse_assets(col(r, "납세_납부액(천원)")),
            "t5": parse_assets(col(r, "납세_5년체납(천원)")),
            "tc": parse_assets(col(r, "납세_현체납(천원)")),
            "c": parse_criminal(col(r, "전과")),
            "cr": col(r, "전과") or "없음",
            "rc": col(r, "입후보횟수"),
            "ph": ph,
            "h": hubo_id,
            "pl": pledges_db.get(hubo_id) or None,   # [{type,url},...] or None
        })
    return out


# ---------- HTML template ---------- #

TEMPLATE = r"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>2026 지방선거 후보자 분석 대시보드</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
  :root {
    --primary: #5645d4;
    --primary-pressed: #4534b3;
    --on-primary: #ffffff;
    --brand-navy: #0a1530;
    --brand-navy-mid: #1a2a52;
    --link-blue: #0075de;
    --tint-peach: #ffe8d4;
    --tint-rose: #fde0ec;
    --tint-mint: #d9f3e1;
    --tint-lavender: #e6e0f5;
    --tint-sky: #dcecfa;
    --tint-yellow: #fef7d6;
    --tint-yellow-bold: #f9e79f;
    --tint-cream: #f8f5e8;
    --tint-gray: #f0eeec;
    --canvas: #ffffff;
    --surface: #f6f5f4;
    --surface-soft: #fafaf9;
    --hairline: #e5e3df;
    --hairline-soft: #ede9e4;
    --hairline-strong: #c8c4be;
    --ink-deep: #000000;
    --ink: #1a1a1a;
    --charcoal: #37352f;
    --slate: #5d5b54;
    --steel: #787671;
    --stone: #a4a097;
    --muted: #bbb8b1;
    --on-dark: #ffffff;
    --on-dark-muted: #a4a097;
    --shadow-mockup: rgba(15, 15, 15, 0.20) 0px 24px 48px -8px;
    --shadow-card: rgba(15, 15, 15, 0.06) 0px 4px 12px 0px;
    --r-sm: 6px; --r-md: 8px; --r-lg: 12px; --r-xl: 16px; --r-full: 9999px;
  }
  * { box-sizing: border-box; }
  html, body { margin: 0; padding: 0; }
  body {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Apple SD Gothic Neo', '맑은 고딕', sans-serif;
    background: var(--canvas);
    color: var(--ink);
    font-size: 16px;
    line-height: 1.55;
    -webkit-font-smoothing: antialiased;
  }

  /* ----- Hero band ----- */
  .hero {
    background: var(--brand-navy);
    color: var(--on-dark);
    padding: 72px 32px 120px;
    position: relative;
    /* z-index 미지정 → stacking context 안 만듦.
       그래야 hero 안의 검색 dropdown이 container를 가로질러 위로 뜰 수 있음. */
  }
  .hero-bg {
    position: absolute; inset: 0; overflow: hidden; pointer-events: none; z-index: 0;
  }
  .hero-bg::before, .hero-bg::after {
    content: ""; position: absolute; border-radius: 50%;
  }
  .hero-bg::before { left: 8%; top: 18%; width: 14px; height: 14px; background: #ff64c8; box-shadow: 280px 80px 0 #f5d75e, 540px -30px 0 #1aae39, 780px 120px 0 #ff7920, 1020px 40px 0 #2a9d99, 80px 220px 0 #7b3ff2; }
  .hero-bg::after { right: 10%; top: 60%; width: 16px; height: 16px; background: #ff7920; box-shadow: -180px -120px 0 #d6b6f6, -420px 60px 0 #0075de, -650px -40px 0 #ff64c8, -900px 100px 0 #1aae39; }
  .hero-inner { max-width: 1200px; margin: 0 auto; text-align: center; position: relative; z-index: 1; }
  .hero .eyebrow {
    font-size: 11px; font-weight: 600; letter-spacing: 1px; text-transform: uppercase;
    color: var(--on-dark-muted); margin-bottom: 12px;
  }
  .hero h1 {
    font-size: 56px; font-weight: 600; line-height: 1.10; letter-spacing: -1px; margin: 0 0 16px;
    color: var(--on-dark);
  }
  .hero p.subtitle {
    font-size: 18px; font-weight: 400; color: var(--on-dark-muted); margin: 0 0 36px;
  }
  /* Search */
  .search-wrap {
    max-width: 580px; margin: 0 auto; position: relative;
    z-index: 100;
  }
  .search-input {
    width: 100%; height: 56px; padding: 0 56px 0 24px;
    border-radius: var(--r-md); border: 1px solid transparent;
    background: var(--canvas); color: var(--ink); font-size: 16px;
    font-family: inherit; box-shadow: var(--shadow-mockup); outline: none;
  }
  .search-input:focus { border-color: var(--primary); border-width: 2px; padding: 0 56px 0 23px; }
  .search-icon {
    position: absolute; right: 18px; top: 50%; transform: translateY(-50%);
    width: 20px; height: 20px; color: var(--steel); pointer-events: none;
  }
  .search-hint { color: var(--on-dark-muted); font-size: 13px; margin-top: 12px; }

  /* Search dropdown */
  .search-results {
    position: fixed;
    /* top/left/width 는 JS 가 입력창 좌표 기반으로 인라인 설정 */
    max-height: min(60vh, 460px); overflow-y: auto;
    background: var(--canvas); border: 1px solid var(--hairline);
    border-radius: var(--r-lg); box-shadow: var(--shadow-mockup);
    text-align: left; z-index: 9999; padding: 6px;
  }
  .search-results.hidden { display: none; }
  .search-result {
    display: grid;
    grid-template-columns: 44px 1fr auto;
    align-items: center; gap: 12px;
    padding: 10px 12px;
    border-radius: var(--r-sm); cursor: pointer; color: var(--ink);
  }
  .search-result + .search-result { border-top: 1px solid var(--hairline-soft); border-radius: 0; }
  .search-result:hover { background: var(--surface); }
  .search-result .thumb {
    width: 44px; height: 56px; border-radius: var(--r-sm); object-fit: cover;
    background: var(--surface); display: block;
  }
  .search-result .meta { min-width: 0; }
  .search-result .name-line { display: flex; align-items: baseline; gap: 6px; }
  .search-result .name {
    font-weight: 600; font-size: 15px; color: var(--ink);
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }
  .search-result .hanja-mini { font-size: 12px; color: var(--steel); font-weight: 400; }
  .search-result .sub {
    font-size: 12px; color: var(--steel); margin-top: 2px;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
  }
  .search-result .party-pill {
    font-size: 11px; font-weight: 600; padding: 3px 10px; border-radius: var(--r-full);
    white-space: nowrap;
  }
  .search-empty {
    padding: 32px 20px; text-align: center; color: var(--steel); font-size: 13px;
  }
  .search-empty .em { font-size: 28px; display: block; margin-bottom: 6px; }

  /* ----- Main area lifted from hero with negative margin ----- */
  .container {
    max-width: 1200px; margin: -72px auto 0; padding: 0 32px 80px;
    position: relative; z-index: 10;
  }

  /* ----- Party tabs ----- */
  .party-tabs {
    background: var(--canvas);
    border: 1px solid var(--hairline);
    border-radius: var(--r-lg);
    padding: 12px;
    margin-bottom: 16px;
    display: flex; flex-wrap: wrap; gap: 6px;
    box-shadow: var(--shadow-card);
  }
  .party-tab {
    border: 1px solid var(--hairline);
    background: transparent;
    color: var(--steel);
    padding: 8px 16px;
    border-radius: var(--r-full);
    font-size: 14px; font-weight: 500; line-height: 1.30;
    cursor: pointer; font-family: inherit;
    display: inline-flex; align-items: center; gap: 6px;
    transition: background 0.12s ease, color 0.12s ease, border-color 0.12s ease;
  }
  .party-tab:hover { background: var(--surface); color: var(--ink); }
  .party-tab.active {
    background: var(--ink-deep); color: var(--on-dark); border-color: var(--ink-deep);
  }
  .party-tab .dot {
    width: 8px; height: 8px; border-radius: 50%; display: inline-block;
  }
  .party-tab .count {
    font-size: 12px; font-weight: 500; color: var(--steel); margin-left: 2px;
  }
  .party-tab.active .count { color: var(--on-dark-muted); }

  /* Secondary filters */
  .filters {
    background: var(--canvas);
    border: 1px solid var(--hairline);
    border-radius: var(--r-lg);
    padding: 16px 20px;
    margin-bottom: 24px;
    display: flex; gap: 24px; align-items: center; flex-wrap: wrap;
  }
  .filter-group { display: flex; flex-direction: column; gap: 6px; }
  .filter-group label {
    font-size: 11px; font-weight: 600; color: var(--steel);
    text-transform: uppercase; letter-spacing: 0.5px;
  }
  .filter-group select {
    height: 36px; padding: 0 12px; border-radius: var(--r-md);
    border: 1px solid var(--hairline-strong); background: var(--canvas);
    color: var(--ink); font-size: 14px; font-family: inherit;
    cursor: pointer; min-width: 180px;
  }
  .filter-count {
    margin-left: auto; font-size: 14px; color: var(--steel); font-weight: 500;
  }
  .filter-count strong { color: var(--ink); font-weight: 600; }

  /* KPI cards */
  .kpis {
    display: grid; grid-template-columns: repeat(5, 1fr); gap: 16px; margin-bottom: 32px;
  }
  @media (max-width: 1100px) { .kpis { grid-template-columns: repeat(2, 1fr); } }
  @media (max-width: 600px) { .kpis { grid-template-columns: 1fr; } }
  .kpi {
    padding: 24px; border-radius: var(--r-lg); color: var(--charcoal);
    background: var(--canvas);
    border: 1px solid var(--hairline);
    position: relative;
    overflow: hidden;
  }
  .kpi::before {
    content: ""; position: absolute; left: 0; top: 0; right: 0; height: 4px;
    background: var(--primary);
  }
  /* 기존 컬러 클래스는 무력화 — 모두 통일 */
  .kpi.peach, .kpi.mint, .kpi.lavender, .kpi.rose, .kpi.sky { background: var(--canvas); }
  .kpi .label {
    font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
    color: var(--steel); margin-bottom: 8px;
  }
  .kpi .value {
    font-size: 32px; font-weight: 600; color: var(--ink); line-height: 1.1; letter-spacing: -0.5px;
  }
  .kpi .value .unit { font-size: 16px; font-weight: 500; margin-left: 4px; color: var(--steel); }
  .kpi .sub { font-size: 12px; color: var(--steel); margin-top: 6px; }

  /* Sections */
  section {
    background: var(--canvas);
    border: 1px solid var(--hairline);
    border-radius: var(--r-lg);
    padding: 28px;
    margin-bottom: 16px;
  }
  section h2 {
    font-size: 22px; font-weight: 600; line-height: 1.30;
    margin: 0 0 6px; color: var(--ink); letter-spacing: -0.2px;
  }
  section .sec-desc {
    font-size: 14px; color: var(--steel); margin-bottom: 20px;
  }
  h3 {
    font-size: 14px; font-weight: 600; color: var(--charcoal); margin: 0 0 12px;
  }
  .filter-tag {
    display: inline-block; vertical-align: middle;
    font-size: 11px; font-weight: 600;
    padding: 2px 8px; border-radius: var(--r-full);
    background: var(--tint-lavender); color: var(--brand-purple-800, #391c57);
    margin-left: 8px; letter-spacing: 0;
  }
  .filter-tag.live {
    background: var(--tint-mint); color: var(--brand-green, #1aae39);
  }
  .chart-block { margin-top: 20px; }
  .chart-block:first-child { margin-top: 0; }
  .chart-wrap { position: relative; height: 360px; }
  .chart-wrap.tall { height: 520px; }
  .chart-wrap.taller { height: 640px; }
  .chart-wrap.short { height: 280px; }
  .chart-wrap.wider { width: 100%; }

  /* Race comparison cards (inside modal) */
  .compare-grid {
    display: grid; grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
    gap: 16px;
  }
  .compare-card {
    background: var(--canvas); border: 1px solid var(--hairline);
    border-radius: var(--r-lg); overflow: hidden; cursor: pointer;
    transition: transform 0.12s ease, box-shadow 0.12s ease;
    display: flex; flex-direction: column;
  }
  .compare-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-card); }
  .compare-card .strip { height: 6px; width: 100%; }
  .compare-card .photo-wrap {
    width: 100%; aspect-ratio: 4/5; background: var(--surface);
    display: flex; align-items: center; justify-content: center;
    color: var(--stone); font-size: 13px;
  }
  .compare-card .photo-wrap img { width: 100%; height: 100%; object-fit: cover; }
  .compare-card .info { padding: 16px; }
  .compare-card .gh-party {
    display: flex; align-items: center; gap: 8px; margin-bottom: 8px;
    font-size: 11px; color: var(--steel);
  }
  .compare-card .gh-num {
    background: var(--ink-deep); color: var(--on-dark);
    width: 22px; height: 22px; border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 12px; font-weight: 600;
  }
  .compare-card .name {
    font-size: 22px; font-weight: 600; color: var(--ink);
    letter-spacing: -0.3px; line-height: 1.2; margin-bottom: 12px;
  }
  .compare-card .name .hanja { font-size: 13px; color: var(--steel); font-weight: 400; }
  .compare-card .stats {
    display: grid; grid-template-columns: 1fr 1fr; gap: 8px;
    padding-top: 12px; border-top: 1px solid var(--hairline-soft);
  }
  .compare-card .stat .k {
    font-size: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
    color: var(--steel); margin-bottom: 2px;
  }
  .compare-card .stat .v {
    font-size: 14px; font-weight: 600; color: var(--ink);
  }
  .compare-card .stat .v.bad { color: #dc2626; }
  .compare-card .stat .v.good { color: var(--brand-green, #1aae39); }

  /* Tables */
  .scroll { max-height: 540px; overflow-y: auto; border: 1px solid var(--hairline); border-radius: var(--r-md); }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  table th {
    background: var(--surface); color: var(--steel); font-weight: 600;
    text-align: left; padding: 10px 12px; font-size: 11px;
    text-transform: uppercase; letter-spacing: 0.5px;
    position: sticky; top: 0; border-bottom: 1px solid var(--hairline);
  }
  table td { padding: 10px 12px; border-bottom: 1px solid var(--hairline-soft); color: var(--charcoal); }
  table tbody tr:last-child td { border-bottom: 0; }
  table tbody tr { cursor: pointer; }
  table tbody tr:hover { background: var(--surface-soft); }
  table td.num { text-align: right; font-variant-numeric: tabular-nums; font-weight: 500; }
  .party-pill {
    display: inline-block; padding: 2px 10px; border-radius: var(--r-full);
    font-size: 11px; font-weight: 600; line-height: 1.4;
  }
  .note { font-size: 12px; color: var(--steel); margin-top: 12px; }

  /* Compare button */
  .btn-compare {
    height: 40px; padding: 0 20px; align-self: end;
    background: var(--primary); color: var(--on-primary);
    border: 0; border-radius: var(--r-md); cursor: pointer;
    font-family: inherit; font-size: 14px; font-weight: 600;
    transition: background 0.12s ease;
  }
  .btn-compare:hover:not(:disabled) { background: var(--primary-pressed); }
  .btn-compare:disabled {
    background: var(--hairline); color: var(--muted); cursor: not-allowed;
  }

  /* ===== Generic modal ===== */
  .modal-overlay {
    position: fixed; inset: 0; background: rgba(15, 15, 15, 0.55);
    z-index: 100; display: flex; align-items: center; justify-content: center;
    padding: 24px;
    backdrop-filter: blur(4px);
  }
  .modal-overlay.hidden { display: none; }
  .modal-card {
    background: var(--canvas); border-radius: var(--r-lg);
    max-width: 760px; width: 100%; max-height: 92vh; overflow-y: auto;
    box-shadow: var(--shadow-mockup); position: relative;
  }
  .modal-card.modal-wide { max-width: 1100px; }
  .modal-close {
    position: absolute; right: 16px; top: 16px;
    width: 32px; height: 32px; border-radius: var(--r-md);
    background: var(--surface); border: 0; cursor: pointer;
    font-size: 14px; color: var(--steel); font-family: inherit;
    z-index: 2;
    display: flex; align-items: center; justify-content: center;
  }
  .modal-close:hover { background: var(--hairline); color: var(--ink); }
  .modal-header-strip {
    padding: 24px 32px 20px;
    border-bottom: 1px solid var(--hairline);
    display: flex; justify-content: space-between; align-items: flex-end;
    gap: 16px;
  }
  .modal-eyebrow {
    font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 1px;
    color: var(--steel); margin-bottom: 4px;
  }
  .modal-title {
    font-size: 24px; font-weight: 600; line-height: 1.20;
    margin: 0; color: var(--ink); letter-spacing: -0.3px;
  }
  .modal-badge {
    flex-shrink: 0;
    background: var(--primary); color: var(--on-primary);
    padding: 6px 12px; border-radius: var(--r-full);
    font-size: 12px; font-weight: 600;
  }
  .modal-body { padding: 24px 32px 32px; }

  /* ===== Detail (person) modal — clean grouped layout ===== */
  .detail-hero {
    padding: 32px 32px 24px;
    background: linear-gradient(180deg, var(--surface-soft) 0%, var(--canvas) 100%);
    display: grid; grid-template-columns: 160px 1fr; gap: 28px;
    border-bottom: 1px solid var(--hairline);
  }
  .detail-photo {
    width: 160px; height: 200px; border-radius: var(--r-lg); object-fit: cover;
    background: var(--surface); border: 1px solid var(--hairline);
    box-shadow: var(--shadow-card);
  }
  .detail-photo-fallback {
    display: flex; align-items: center; justify-content: center;
    color: var(--stone); font-size: 13px; text-align: center;
  }
  .detail-meta { min-width: 0; align-self: center; }
  .detail-meta .badges { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 12px; }
  .detail-badge {
    font-size: 11px; font-weight: 600; padding: 4px 10px;
    border-radius: var(--r-full); background: var(--canvas);
    border: 1px solid var(--hairline); color: var(--charcoal);
  }
  .detail-badge.gh {
    background: var(--ink-deep); color: var(--on-dark); border-color: var(--ink-deep);
  }
  .detail-name {
    font-size: 36px; font-weight: 600; line-height: 1.15; letter-spacing: -0.5px;
    margin: 0 0 4px; color: var(--ink);
  }
  .detail-name .hanja { font-size: 16px; color: var(--steel); font-weight: 400; margin-left: 8px; }
  .detail-sub {
    font-size: 14px; color: var(--steel); margin-bottom: 2px;
  }
  .detail-sub strong { color: var(--charcoal); font-weight: 500; }

  /* Stats strip */
  .detail-stats {
    display: grid; grid-template-columns: repeat(4, 1fr);
    border-bottom: 1px solid var(--hairline);
    background: var(--canvas);
  }
  @media (max-width: 600px) { .detail-stats { grid-template-columns: repeat(2, 1fr); } }
  .detail-stat {
    padding: 18px 20px;
    border-right: 1px solid var(--hairline-soft);
  }
  .detail-stat:last-child { border-right: 0; }
  .detail-stat .k {
    font-size: 10px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
    color: var(--steel); margin-bottom: 4px;
  }
  .detail-stat .v {
    font-size: 20px; font-weight: 600; color: var(--ink); letter-spacing: -0.3px;
  }
  .detail-stat .v.bad { color: #dc2626; }
  .detail-stat .v.good { color: var(--brand-green, #1aae39); }

  /* Body — grouped */
  .detail-body { padding: 24px 32px 32px; }
  .detail-group {
    border: 1px solid var(--hairline);
    border-radius: var(--r-lg);
    padding: 4px 0;
    margin-bottom: 12px;
    overflow: hidden;
  }
  .detail-group-title {
    font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;
    color: var(--steel);
    padding: 10px 18px 6px;
    background: var(--surface-soft);
    border-bottom: 1px solid var(--hairline);
    margin: -4px 0 0;
  }
  .detail-row {
    display: grid; grid-template-columns: 120px 1fr;
    gap: 16px; padding: 12px 18px;
    border-bottom: 1px solid var(--hairline-soft);
  }
  .detail-row:last-child { border-bottom: 0; }
  .detail-row .k {
    font-size: 12px; color: var(--steel); font-weight: 500;
    align-self: start; padding-top: 1px;
  }
  .detail-row .v {
    font-size: 14px; color: var(--charcoal);
    white-space: pre-wrap; line-height: 1.55;
  }
  .detail-row .v.num { font-variant-numeric: tabular-nums; font-weight: 500; }

  .detail-fullpage {
    display: flex; align-items: center; justify-content: space-between;
    margin: 0 32px 32px; padding: 14px 20px;
    background: var(--primary); color: var(--on-primary);
    border-radius: var(--r-md);
    font-size: 14px; font-weight: 600; text-decoration: none;
    transition: background 0.12s ease;
  }
  .detail-fullpage:hover { background: var(--primary-pressed); }
  .detail-fullpage .arrow { font-size: 18px; }

  /* Pledge download buttons in modal */
  .pledge-dl-row { display: flex; gap: 10px; flex-wrap: wrap; }
  .pledge-dl {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 8px 14px; background: var(--surface); border: 1px solid var(--hairline);
    border-radius: var(--r-md); color: var(--charcoal);
    font-size: 13px; font-weight: 600; text-decoration: none;
    transition: border-color 0.12s ease, color 0.12s ease;
  }
  .pledge-dl:hover { border-color: var(--primary); color: var(--primary); text-decoration: none; }
  .pledge-dl .ico { color: var(--primary); }

  /* Q&A inside modal */
  .qa-intro { font-size: 12px; color: var(--steel); margin-bottom: 10px; line-height: 1.55; }
  .qa-suggestions { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 10px; }
  .qa-chip {
    padding: 5px 11px; background: var(--canvas); border: 1px solid var(--hairline);
    border-radius: var(--r-full); font-size: 12px; color: var(--charcoal);
    font-family: inherit; cursor: pointer;
  }
  .qa-chip:hover { background: var(--surface-soft); border-color: var(--primary); }
  .qa-msgs {
    display: flex; flex-direction: column; gap: 8px;
    margin-bottom: 12px; max-height: 360px; overflow-y: auto; padding-right: 4px;
  }
  .qa-msg {
    padding: 10px 14px; border-radius: var(--r-md); font-size: 13px; line-height: 1.6;
    white-space: pre-wrap; word-break: break-word;
  }
  .qa-msg.user { background: var(--primary); color: var(--on-primary); align-self: flex-end; max-width: 80%; }
  .qa-msg.bot { background: var(--surface); color: var(--charcoal); align-self: flex-start; max-width: 90%; border: 1px solid var(--hairline); }
  .qa-msg.bot.error { background: #fff1f0; border-color: #fecaca; color: #991b1b; }
  .qa-msg.bot.loading { color: var(--steel); font-style: italic; }
  .qa-input-row { display: flex; gap: 8px; }
  .qa-input-row input {
    flex: 1; padding: 11px 14px; border: 1px solid var(--hairline);
    border-radius: var(--r-md); font-family: inherit; font-size: 14px;
  }
  .qa-input-row input:focus { outline: none; border-color: var(--primary); }
  .qa-send {
    padding: 11px 18px; background: var(--primary); color: var(--on-primary);
    border: none; border-radius: var(--r-md); font-weight: 600;
    cursor: pointer; font-family: inherit; font-size: 14px;
  }
  .qa-send:hover:not(:disabled) { background: var(--primary-pressed); }
  .qa-send:disabled { opacity: 0.5; cursor: not-allowed; }

  /* Floating Chatbot */
  .chat-fab {
    position: fixed; right: 24px; bottom: 24px; z-index: 90;
    display: flex; align-items: center; gap: 8px;
    padding: 14px 20px;
    background: var(--primary); color: var(--on-primary);
    border: none; border-radius: var(--r-full);
    font-family: inherit; font-size: 14px; font-weight: 600;
    cursor: pointer;
    box-shadow: 0 8px 24px rgba(86, 69, 212, 0.35);
    transition: transform 0.12s ease, box-shadow 0.12s ease;
  }
  .chat-fab:hover { transform: translateY(-2px); box-shadow: 0 12px 28px rgba(86, 69, 212, 0.42); }
  .chat-fab:active { transform: translateY(0); }
  .chat-fab .ico { font-size: 18px; }
  .chat-fab.active { display: none; }

  .chat-panel {
    position: fixed; right: 24px; bottom: 24px; z-index: 95;
    width: 400px; max-width: calc(100vw - 32px);
    height: 580px; max-height: calc(100vh - 48px);
    background: var(--canvas);
    border: 1px solid var(--hairline);
    border-radius: var(--r-lg);
    box-shadow: 0 20px 60px rgba(15, 15, 15, 0.18);
    display: flex; flex-direction: column;
    overflow: hidden;
  }
  .chat-panel.hidden { display: none; }
  .chat-header {
    padding: 16px 20px;
    background: var(--brand-navy); color: var(--on-dark);
    display: flex; align-items: center; justify-content: space-between;
    flex-shrink: 0;
  }
  .chat-title { font-size: 15px; font-weight: 700; letter-spacing: -0.01em; }
  .chat-subtitle { font-size: 11px; color: var(--on-dark-muted); margin-top: 2px; }
  .chat-close {
    background: transparent; border: none; color: var(--on-dark);
    font-size: 18px; cursor: pointer; padding: 4px 8px; border-radius: var(--r-sm);
  }
  .chat-close:hover { background: rgba(255,255,255,0.1); }
  .chat-body {
    flex: 1; overflow-y: auto;
    padding: 16px 18px; background: var(--surface-soft);
  }
  .chat-intro {
    font-size: 12px; color: var(--steel); line-height: 1.65;
    background: var(--canvas); border: 1px solid var(--hairline);
    border-radius: var(--r-md); padding: 10px 14px;
    margin-bottom: 12px;
  }
  .chat-msgs { display: flex; flex-direction: column; gap: 8px; padding-bottom: 12px; }
  .chat-input-row {
    display: flex; gap: 8px; padding: 12px 14px;
    border-top: 1px solid var(--hairline); background: var(--canvas);
    flex-shrink: 0;
  }
  .chat-input-row input {
    flex: 1; padding: 10px 14px;
    border: 1px solid var(--hairline); border-radius: var(--r-md);
    font-family: inherit; font-size: 14px;
  }
  .chat-input-row input:focus { outline: none; border-color: var(--primary); }

  @media (max-width: 600px) {
    .chat-fab { padding: 12px 16px; font-size: 13px; right: 16px; bottom: 16px; }
    .chat-panel { right: 8px; bottom: 8px; left: 8px; width: auto; height: calc(100vh - 16px); max-height: 720px; }
  }

  .empty-state {
    padding: 60px 20px; text-align: center; color: var(--steel);
  }
  .empty-state .em { font-size: 32px; display: block; margin-bottom: 8px; }
</style>
</head>
<body>

<!-- HERO -->
<header class="hero">
  <div class="hero-bg"></div>
  <div class="hero-inner">
    <div class="eyebrow">중앙선거관리위원회 · info.nec.go.kr</div>
    <h1>2026 지방선거 후보자 분석</h1>
    <p class="subtitle">제9회 전국동시지방선거 · 2026년 6월 3일 · 후보자 7,826명</p>
    <div class="search-wrap">
      <input id="search" class="search-input" type="text" placeholder="후보자 이름을 검색하세요 (예: 정원오)" autocomplete="off">
      <svg class="search-icon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.3-4.3"/></svg>
    </div>
    <div class="search-hint">선택한 후보자의 사진 · 학력 · 경력 · 재산 · 전과 정보를 확인할 수 있습니다</div>
  </div>
</header>

<div class="container">
  <!-- PARTY TABS -->
  <div id="party-tabs" class="party-tabs"></div>

  <!-- SECONDARY FILTERS -->
  <div class="filters">
    <div class="filter-group">
      <label for="f-election">선거 종류</label>
      <select id="f-election"></select>
    </div>
    <div class="filter-group">
      <label for="f-sido">지역</label>
      <select id="f-sido"></select>
    </div>
    <div class="filter-group">
      <label for="f-sgg">선거구</label>
      <select id="f-sgg" disabled></select>
    </div>
    <button id="btn-compare" class="btn-compare" disabled>비교하기</button>
    <div class="filter-count" id="count-display"></div>
  </div>

  <!-- KPI -->
  <div class="kpis">
    <div class="kpi peach"><div class="label">총 후보자</div><div class="value"><span id="kpi-count"></span><span class="unit">명</span></div><div class="sub" id="kpi-count-sub"></div></div>
    <div class="kpi mint"><div class="label">평균 나이</div><div class="value"><span id="kpi-age"></span><span class="unit">세</span></div><div class="sub" id="kpi-age-sub"></div></div>
    <div class="kpi lavender"><div class="label">평균 자산</div><div class="value" id="kpi-assets"></div><div class="sub" id="kpi-assets-sub"></div></div>
    <div class="kpi rose"><div class="label">전과 보유율</div><div class="value"><span id="kpi-crim"></span><span class="unit">%</span></div><div class="sub" id="kpi-crim-sub"></div></div>
    <div class="kpi sky"><div class="label">남성 병역 미필률</div><div class="value"><span id="kpi-mil"></span><span class="unit">%</span></div><div class="sub">군 면제·미필 합산</div></div>
  </div>

  <section id="sec-party">
    <h2>정당별 후보자 분포</h2>
    <p class="sec-desc">한 정당당 등록된 후보자 수 (단위: 명)</p>
    <div class="chart-wrap taller"><canvas id="ch-party"></canvas></div>
  </section>

  <section>
    <h2>전과 분석</h2>
    <p class="sec-desc">후보자가 신고한 전과 기록 (벌금형 이상)</p>
    <div class="chart-block">
      <h3>정당별 전과 보유율 <span class="filter-tag">전체 기준 · 필터 무관</span> <span style="font-weight:400;color:var(--steel)">(단위: %)</span></h3>
      <div class="chart-wrap tall"><canvas id="ch-crim-party"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>지역별 전과 보유율 <span class="filter-tag live">필터 반영</span> <span style="font-weight:400;color:var(--steel)">(단위: %)</span></h3>
      <div class="chart-wrap tall"><canvas id="ch-crim-sido"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>전과 많은 후보자 Top 30</h3>
      <div class="scroll">
        <table id="tbl-crim-top"><thead><tr>
          <th>#</th><th>이름</th><th>정당</th><th>지역</th><th>선거</th><th class="num">건수</th>
        </tr></thead><tbody></tbody></table>
      </div>
    </div>
  </section>

  <section>
    <h2>자산 분석</h2>
    <p class="sec-desc">후보자 신고 자산 (천원 단위, 가족 포함) · 미신고 제외</p>
    <div class="chart-block">
      <h3>정당별 평균 자산 <span class="filter-tag">전체 기준 · 필터 무관</span> <span style="font-weight:400;color:var(--steel)">(단위: 원)</span></h3>
      <div class="chart-wrap tall"><canvas id="ch-asset-party"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>지역별 평균 자산 <span class="filter-tag live">필터 반영</span> <span style="font-weight:400;color:var(--steel)">(단위: 원)</span></h3>
      <div class="chart-wrap tall"><canvas id="ch-asset-sido"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>자산 구간별 후보자 수 <span class="filter-tag live">필터 반영</span> <span style="font-weight:400;color:var(--steel)">(단위: 명)</span></h3>
      <div class="chart-wrap"><canvas id="ch-asset-hist"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>부자 Top 30</h3>
      <div class="scroll"><table id="tbl-rich"><thead><tr>
        <th>#</th><th>이름</th><th>정당</th><th>지역</th><th class="num">자산</th>
      </tr></thead><tbody></tbody></table></div>
    </div>
    <div class="chart-block">
      <h3>가난 Top 30 (신고액 낮은 순)</h3>
      <div class="scroll"><table id="tbl-poor"><thead><tr>
        <th>#</th><th>이름</th><th>정당</th><th>지역</th><th class="num">자산</th>
      </tr></thead><tbody></tbody></table></div>
    </div>
  </section>

  <section>
    <h2>나이 분석</h2>
    <p class="sec-desc">후보자 연령 분포 (선거일 기준)</p>
    <div class="chart-block">
      <h3>정당별 평균 나이 <span class="filter-tag">전체 기준 · 필터 무관</span> <span style="font-weight:400;color:var(--steel)">(단위: 세)</span></h3>
      <div class="chart-wrap tall"><canvas id="ch-age-party"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>지역별 평균 나이 <span class="filter-tag live">필터 반영</span> <span style="font-weight:400;color:var(--steel)">(단위: 세)</span></h3>
      <div class="chart-wrap tall"><canvas id="ch-age-sido"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>연령대별 후보자 수 <span class="filter-tag live">필터 반영</span> <span style="font-weight:400;color:var(--steel)">(단위: 명)</span></h3>
      <div class="chart-wrap"><canvas id="ch-age-hist"></canvas></div>
    </div>
  </section>

  <section>
    <h2>병역 · 직업</h2>
    <p class="sec-desc">병역은 남성 후보자 기준 · 직업은 14개 카테고리로 자동 분류</p>
    <div class="chart-block">
      <h3>병역 분포 (단위: 명)</h3>
      <div class="chart-wrap tall"><canvas id="ch-mil"></canvas></div>
    </div>
    <div class="chart-block">
      <h3>직업 분포 (단위: 명)</h3>
      <div class="chart-wrap tall"><canvas id="ch-job"></canvas></div>
    </div>
  </section>
</div>

<!-- SEARCH RESULTS DROPDOWN (body-level fixed so it's never covered) -->
<div id="search-results" class="search-results hidden"></div>

<!-- COMPARISON MODAL -->
<div id="compare-overlay" class="modal-overlay hidden">
  <div class="modal-card modal-wide">
    <button class="modal-close" id="compare-close" aria-label="닫기">✕</button>
    <div class="modal-header-strip">
      <div>
        <div class="modal-eyebrow" id="compare-eyebrow"></div>
        <h2 class="modal-title" id="compare-title"></h2>
      </div>
      <span class="modal-badge" id="compare-count"></span>
    </div>
    <div class="modal-body">
      <div class="compare-grid" id="compare-grid"></div>
    </div>
  </div>
</div>

<!-- DETAIL MODAL -->
<div id="detail-overlay" class="modal-overlay hidden">
  <div class="modal-card" id="detail-card">
    <button class="modal-close" id="detail-close" aria-label="닫기">✕</button>
    <div id="detail-content"></div>
  </div>
</div>

<!-- FLOATING CHATBOT -->
<button id="chatFab" class="chat-fab" aria-label="AI 공약 도우미 열기">
  <span class="ico" aria-hidden="true">💬</span>
  <span class="label">공약 물어보기</span>
</button>
<div id="chatPanel" class="chat-panel hidden" role="dialog" aria-label="AI 공약 도우미">
  <div class="chat-header">
    <div>
      <div class="chat-title">AI 공약 도우미</div>
      <div class="chat-subtitle">2026 지방선거 후보·공약 질문</div>
    </div>
    <button id="chatClose" class="chat-close" aria-label="닫기">✕</button>
  </div>
  <div class="chat-body">
    <div class="chat-intro">
      후보 이름과 함께 질문해 주세요. 예시:<br>
      · "정원오 후보의 청년 공약은?"<br>
      · "김진태 후보 주거 정책 알려줘"
    </div>
    <div class="qa-suggestions" id="chatSuggestions"></div>
    <div class="chat-msgs" id="chatMsgs"></div>
  </div>
  <form id="chatForm" class="chat-input-row">
    <input id="chatInput" type="text" placeholder="후보 이름과 함께 질문…" autocomplete="off">
    <button id="chatSend" class="qa-send" type="submit">전송</button>
  </form>
</div>

<script>
const DATA = __DATA__;
const SIDO_ORDER = ["서울특별시","경기도","인천광역시","부산광역시","대구광역시","광주광역시","대전광역시","울산광역시","세종특별자치시","강원특별자치도","충청북도","충청남도","전북특별자치도","전라남도","경상북도","경상남도","제주특별자치도","전남광주통합특별시"];

const PARTY_COLORS = {
  "더불어민주당": "#004EA2",
  "국민의힘": "#E61E2B",
  "조국혁신당": "#0073CF",
  "개혁신당": "#FF7920",
  "진보당": "#D6001C",
  "정의당": "#FFCC00",
  "기본소득당": "#00B5A8",
  "새미래민주당": "#1AA75E",
  "자유와혁신": "#8A2BE2",
  "노동당": "#EB121A",
  "자유통일당": "#A6324D",
  "무소속": "#9ca3af",
};
const FALLBACK_COLORS = ["#7c3aed","#0891b2","#65a30d","#db2777","#0284c7","#a16207","#16a34a","#9333ea","#d97706","#475569"];
const colorFor = (p, i=0) => PARTY_COLORS[p] || FALLBACK_COLORS[i % FALLBACK_COLORS.length];
// 정당 ↔ 핀 컬러 (밝은 배경)
const pillStyle = (p) => {
  const c = colorFor(p);
  return `background:${c}22;color:${c};`;
};
// 차트의 막대 색을 더 진하게 만들려면 그대로 colorFor 사용

const fmtKr = (n) => {
  if (n === null || n === undefined || Number.isNaN(n)) return "-";
  const abs = Math.abs(n);
  if (abs >= 1e8) return (n/1e8).toFixed(1) + "억";
  if (abs >= 1e4) return (n/1e4).toFixed(1) + "만";
  return n.toLocaleString();
};
const fmtAsset = (cw) => {
  if (cw === null || cw === undefined || Number.isNaN(cw)) return "-";
  const won = cw * 1000;
  const abs = Math.abs(won);
  if (abs >= 1e12) return (won/1e12).toFixed(2) + "조";
  if (abs >= 1e8) return (won/1e8).toFixed(2) + "억";
  if (abs >= 1e4) return (won/1e4).toFixed(0) + "만";
  return won.toLocaleString() + "원";
};
const fmtPct = (n, d=1) => (n*100).toFixed(d) + "%";

// --- party tabs --- //
const partyCounts = {};
for (const x of DATA) partyCounts[x.p] = (partyCounts[x.p] || 0) + 1;
const TOP_PARTIES = Object.entries(partyCounts).sort((a,b)=>b[1]-a[1]).slice(0, 10).map(x=>x[0]);

let activePartyTab = "전체";
const partyTabBar = document.getElementById("party-tabs");
function renderPartyTabs() {
  const items = [
    { name: "전체", color: "#1a1a1a", count: DATA.length },
    ...TOP_PARTIES.map(p => ({ name: p, color: colorFor(p), count: partyCounts[p] })),
  ];
  partyTabBar.innerHTML = items.map(it => `
    <button class="party-tab ${activePartyTab === it.name ? 'active' : ''}" data-party="${it.name}">
      ${it.name !== "전체" ? `<span class="dot" style="background:${it.color}"></span>` : ''}
      ${it.name}
      <span class="count">${it.count.toLocaleString()}</span>
    </button>
  `).join("");
  for (const btn of partyTabBar.querySelectorAll(".party-tab")) {
    btn.addEventListener("click", () => {
      activePartyTab = btn.dataset.party;
      renderPartyTabs();
      render();
    });
  }
}
renderPartyTabs();

// --- secondary filters --- //
const fEl = document.getElementById("f-election");
const fSi = document.getElementById("f-sido");
const fSg = document.getElementById("f-sgg");

function uniqSorted(arr, order) {
  const set = new Set(arr.filter(Boolean));
  if (order) {
    const inOrder = order.filter(x => set.has(x));
    for (const x of set) if (!inOrder.includes(x)) inOrder.push(x);
    return inOrder;
  }
  return [...set].sort();
}
function populateSelect(sel, values, label) {
  sel.innerHTML = `<option value="">${label} (전체)</option>` + values.map(v => `<option>${v}</option>`).join("");
}
populateSelect(fEl, uniqSorted(DATA.map(d => d.e)).reverse(), "선거종류");
populateSelect(fSi, uniqSorted(DATA.map(d => d.s), SIDO_ORDER), "지역");
populateSelect(fSg, [], "선거구");
fSg.disabled = true;

const btnCompare = document.getElementById("btn-compare");
function refreshSggOptions() {
  const e = fEl.value, s = fSi.value;
  if (!e || !s) {
    populateSelect(fSg, [], "선거구");
    fSg.disabled = true;
  } else {
    const opts = uniqSorted(
      DATA.filter(d => d.e === e && d.s === s && d.g).map(d => d.g)
    );
    if (opts.length === 0) {
      populateSelect(fSg, [], "선거구 없음");
      fSg.disabled = true;
    } else {
      populateSelect(fSg, opts, "선거구");
      fSg.disabled = false;
    }
  }
  // 비교하기 버튼: 선거종류 + 지역 둘 다 선택돼야 활성
  btnCompare.disabled = !(e && s);
}
fEl.addEventListener("change", () => { refreshSggOptions(); render(); });
fSi.addEventListener("change", () => { refreshSggOptions(); render(); });
fSg.addEventListener("change", render);
btnCompare.addEventListener("click", openCompareModal);

function currentFilter() {
  const e = fEl.value, s = fSi.value, g = fSg.value;
  return DATA.filter(d =>
    (activePartyTab === "전체" || d.p === activePartyTab) &&
    (!e || d.e === e) &&
    (!s || d.s === s) &&
    (!g || d.g === g)
  );
}

// --- helpers --- //
function avg(arr) { const xs = arr.filter(x => x !== null && x !== undefined); return xs.length ? xs.reduce((a,b)=>a+b,0)/xs.length : null; }
function median(arr) { const xs = arr.filter(x => x !== null && x !== undefined).sort((a,b)=>a-b); return xs.length ? xs[Math.floor(xs.length/2)] : null; }

// --- charts --- //
let charts = {};
function destroyAll() { for (const c of Object.values(charts)) c.destroy(); charts = {}; }

const COMMON_GRID = { color: "#ede9e4" };
const COMMON_TICK = { color: "#5d5b54", font: { family: "Inter, sans-serif", size: 11 } };

function makeBar(elId, labels, values, colors, opts={}) {
  const ctx = document.getElementById(elId);
  const horizontal = opts.horizontal !== false; // default 수평
  const unit = opts.unit || "";

  // value axis tick formatter
  const valueTick = function(v) {
    if (opts.valueTickFmt) return opts.valueTickFmt(v);
    return v.toLocaleString() + unit;
  };
  // category axis: always show label (no autoSkip)
  const categoryAxis = {
    type: "category",
    grid: { display: false },
    ticks: {
      ...COMMON_TICK,
      autoSkip: false,
      maxRotation: horizontal ? 0 : 35,
      minRotation: 0,
      callback: function(v, i) { return this.getLabelForValue ? this.getLabelForValue(v) : labels[i]; },
    }
  };
  const valueAxis = {
    type: "linear",
    beginAtZero: true,
    grid: COMMON_GRID,
    ticks: { ...COMMON_TICK, callback: valueTick },
    title: opts.unit ? { display: true, text: `(단위: ${opts.unit})`, color: "#787671", font: { size: 11 } } : undefined,
  };

  charts[elId] = new Chart(ctx, {
    type: "bar",
    data: { labels, datasets: [{
      data: values,
      backgroundColor: colors || labels.map((_,i)=>FALLBACK_COLORS[i % FALLBACK_COLORS.length]),
      borderRadius: 4,
      borderSkipped: false,
      barPercentage: 0.78,
      categoryPercentage: 0.78,
    }] },
    options: {
      responsive: true, maintainAspectRatio: false,
      indexAxis: horizontal ? "y" : "x",
      layout: { padding: { left: 4, right: 16, top: 4, bottom: 4 } },
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: (ctx) => {
              const v = ctx.parsed[horizontal ? 'x' : 'y'];
              return opts.tooltipFmt ? opts.tooltipFmt(v) : (v.toLocaleString() + unit);
            }
          }
        }
      },
      scales: horizontal
        ? { x: valueAxis, y: categoryAxis }
        : { x: categoryAxis, y: valueAxis }
    }
  });
}

function makeDoughnut(elId, labels, values, colors) {
  const ctx = document.getElementById(elId);
  const total = values.reduce((a,b)=>a+b,0);
  charts[elId] = new Chart(ctx, {
    type: "doughnut",
    data: { labels, datasets: [{ data: values, backgroundColor: colors, borderWidth: 1, borderColor: "#ffffff" }] },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { position: "right", labels: { color: "#37352f", font: { family: "Inter, sans-serif", size: 12 } } },
        tooltip: { callbacks: { label: (ctx) => `${ctx.label}: ${ctx.parsed.toLocaleString()}명 (${(ctx.parsed/total*100).toFixed(1)}%)` } }
      }
    }
  });
}

function fillTable(tblId, rows, render) {
  const tb = document.querySelector(`#${tblId} tbody`);
  if (!rows.length) {
    tb.innerHTML = `<tr><td colspan="6" class="empty-state">데이터 없음</td></tr>`;
    return;
  }
  tb.innerHTML = rows.map((r,i) => `<tr data-hubo="${r.h}">${render(r,i)}</tr>`).join("");
  for (const tr of tb.querySelectorAll("tr")) {
    tr.addEventListener("click", () => openDetail(tr.dataset.hubo));
  }
}

// --- detail (person) modal --- //
const detailOverlay = document.getElementById("detail-overlay");
const detailContent = document.getElementById("detail-content");
document.getElementById("detail-close").addEventListener("click", () => detailOverlay.classList.add("hidden"));
detailOverlay.addEventListener("click", (e) => { if (e.target === detailOverlay) detailOverlay.classList.add("hidden"); });
document.addEventListener("keydown", (e) => {
  if (e.key === "Escape") {
    detailOverlay.classList.add("hidden");
    compareOverlay.classList.add("hidden");
  }
});

function thumbUrl(ph) {
  if (!ph) return "";
  // 로컬 photos/*.jpg 는 이미 썸네일 이미지 (다운로드 시 thumbnail 받음) — 변환 불필요
  if (!ph.startsWith("http")) return ph;
  return ph.replace(/\/gicho\/(\d+)\.JPG$/i, "/gicho/thumbnail.$1.JPG");
}
// 200 OK + HTML 응답을 잡아내는 견고한 img 로더 (attribute 안전 escape)
function htmlAttrEscape(s) {
  return String(s).replace(/&/g, "&amp;").replace(/"/g, "&quot;").replace(/'/g, "&#39;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
}
function safeImg(url, cls, alt, fallback) {
  if (!url) return fallback;
  const onload = "if(this.naturalWidth<10){this.onload=null;this.dispatchEvent(new Event('error'));}";
  const onerrorJs = "this.outerHTML=" + JSON.stringify(fallback);
  return `<img class="${htmlAttrEscape(cls)}" src="${htmlAttrEscape(url)}" alt="${htmlAttrEscape(alt)}" onload="${htmlAttrEscape(onload)}" onerror="${htmlAttrEscape(onerrorJs)}">`;
}

function detailRow(k, v, cls) {
  if (v === null || v === undefined || v === '') return '';
  const safeV = String(v).replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
  return `<div class="detail-row"><div class="k">${k}</div><div class="v ${cls || ''}">${safeV}</div></div>`;
}
function detailGroup(title, rows) {
  const inner = rows.filter(Boolean).join("");
  if (!inner) return '';
  return `<div class="detail-group"><div class="detail-group-title">${title}</div>${inner}</div>`;
}

function openDetail(huboId) {
  const r = DATA.find(d => d.h == huboId);
  if (!r) return;
  const photoHtml = safeImg(
    r.ph, "detail-photo", `${r.n} 사진`,
    `<div class="detail-photo detail-photo-fallback">사진<br>없음</div>`
  );
  const crimCls = r.c > 0 ? 'bad' : 'good';
  detailContent.innerHTML = `
    <div class="detail-hero">
      ${photoHtml}
      <div class="detail-meta">
        <div class="badges">
          ${r.gh ? `<span class="detail-badge gh">기호 ${r.gh}</span>` : ''}
          <span class="detail-badge" style="${pillStyle(r.p)}">${r.p}</span>
          <span class="detail-badge">${r.e}</span>
        </div>
        <h2 class="detail-name">${r.n}${r.hj ? `<span class="hanja">${r.hj}</span>` : ''}</h2>
        <div class="detail-sub"><strong>${r.s}</strong>${r.g ? ' · ' + r.g : ''}</div>
        <div class="detail-sub">${r.gd || ''}${r.bd ? ' · ' + r.bd : ''}${r.a ? ' (만 ' + r.a + '세)' : ''}</div>
      </div>
    </div>
    <div class="detail-stats">
      <div class="detail-stat"><div class="k">나이</div><div class="v">${r.a || '-'}세</div></div>
      <div class="detail-stat"><div class="k">자산</div><div class="v">${fmtAsset(r.as)}</div></div>
      <div class="detail-stat"><div class="k">전과</div><div class="v ${crimCls}">${r.c > 0 ? r.c + '건' : '없음'}</div></div>
      <div class="detail-stat"><div class="k">병역</div><div class="v">${r.gd === '남' ? r.m : '-'}</div></div>
    </div>
    <div class="detail-body">
      ${detailGroup('기본 정보', [
        detailRow('주소', r.addr),
        detailRow('직업', r.jr),
      ])}
      ${detailGroup('학력 · 경력', [
        detailRow('학력', r.edu),
        detailRow('경력', r.car),
        detailRow('입후보 횟수', r.rc),
      ])}
      ${detailGroup('재산 · 납세 (단위: 천원)', [
        detailRow('재산신고액', r.as !== null ? r.as.toLocaleString() + '  (' + fmtAsset(r.as) + ')' : '미신고', 'num'),
        detailRow('납세 납부액', r.tp !== null ? r.tp.toLocaleString() : '-', 'num'),
        detailRow('5년 체납액', r.t5 !== null ? r.t5.toLocaleString() : '-', 'num'),
        detailRow('현재 체납액', r.tc !== null ? r.tc.toLocaleString() : '-', 'num'),
      ])}
      ${detailGroup('병역 · 전과', [
        detailRow('병역', r.mr),
        detailRow('전과', r.cr),
      ])}
      ${renderPledgeBlock(r)}
      ${renderQaBlock(r)}
    </div>
    <a class="detail-fullpage" href="candidates/${r.h}.html">
      <span>전체 공약 PDF 미리보기</span>
      <span class="arrow">→</span>
    </a>`;
  detailOverlay.classList.remove("hidden");
  // Q&A 폼 와이어업 — 모달 매번 열릴 때마다
  wireQaForm(r);
}

function renderPledgeBlock(r) {
  if (!r.pl || !r.pl.length) {
    return `<div class="detail-group">
      <div class="detail-group-title">공약 자료</div>
      <div class="detail-row"><div class="v" style="color:var(--muted)">NEC에 제출된 공약 자료가 없습니다.</div></div>
    </div>`;
  }
  const links = r.pl.map(p =>
    `<a class="pledge-dl" href="${p.url}" download="${r.n}_${p.type}.pdf" target="_blank" rel="noopener">
       <span class="ico">↓</span>${p.type} PDF 다운로드
     </a>`
  ).join('');
  return `<div class="detail-group">
    <div class="detail-group-title">공약 자료</div>
    <div class="detail-row" style="grid-template-columns:1fr">
      <div class="v"><div class="pledge-dl-row">${links}</div></div>
    </div>
  </div>`;
}

function renderQaBlock(r) {
  if (!r.pl || !r.pl.length) {
    return `<div class="detail-group">
      <div class="detail-group-title">AI 공약 Q&A</div>
      <div class="detail-row"><div class="v" style="color:var(--muted)">공약 자료가 없어 Q&A를 사용할 수 없습니다.</div></div>
    </div>`;
  }
  const chips = ['이 후보의 핵심 공약 3가지는?','청년·일자리 공약은?','주거·교통 공약은?','재원 조달 방안이 있어?']
    .map(s => `<button type="button" class="qa-chip">${s}</button>`).join('');
  return `<div class="detail-group">
    <div class="detail-group-title">AI 공약 Q&A · Gemini</div>
    <div style="padding:14px 18px 18px">
      <div class="qa-intro">후보자가 제출한 공약 자료를 바탕으로 답변. 사실관계는 PDF 원문을 직접 확인해 주세요.</div>
      <div class="qa-suggestions">${chips}</div>
      <div class="qa-msgs" id="modalQaMsgs"></div>
      <form id="modalQaForm" class="qa-input-row">
        <input id="modalQaInput" type="text" placeholder="이 후보 공약 중 궁금한 점을 물어보세요…" autocomplete="off">
        <button id="modalQaSend" class="qa-send" type="submit">질문</button>
      </form>
    </div>
  </div>`;
}

function wireQaForm(r) {
  const form = document.getElementById('modalQaForm');
  if (!form) return;
  const ask = async (question) => {
    const msgs = document.getElementById('modalQaMsgs');
    const sendBtn = document.getElementById('modalQaSend');
    const input = document.getElementById('modalQaInput');
    const addMsg = (role, text, extraClass='') => {
      const div = document.createElement('div');
      div.className = 'qa-msg ' + role + (extraClass ? ' '+extraClass : '');
      div.textContent = text;
      msgs.appendChild(div);
      msgs.scrollTop = msgs.scrollHeight;
      return div;
    };
    addMsg('user', question);
    const loading = addMsg('bot', '공약 자료를 읽고 답변 작성 중…', 'loading');
    input.value = '';
    sendBtn.disabled = true;
    try {
      const resp = await fetch('/api/qa', {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({
          question,
          candidate: { huboId: r.h, name: r.n, party: r.p, giho: r.gh, subSgName: r.e, sgg: r.g, sido: r.s },
          pledgeText: '',
          pdfUrls: r.pl || []
        })
      });
      const j = await resp.json().catch(()=>({error:'응답 파싱 실패'}));
      loading.remove();
      if (!resp.ok) { addMsg('bot', j.error || ('서버 오류 '+resp.status), 'error'); return; }
      addMsg('bot', (j.answer || '').trim() || '(빈 응답)');
    } catch(e) {
      loading.remove();
      addMsg('bot', '요청 실패: '+(e.message||e), 'error');
    } finally {
      sendBtn.disabled = false;
    }
  };
  form.addEventListener('submit', e => {
    e.preventDefault();
    const q = document.getElementById('modalQaInput').value.trim();
    if (q) ask(q);
  });
  document.querySelectorAll('.qa-chip').forEach(chip => {
    chip.addEventListener('click', () => ask(chip.textContent));
  });
}

// --- search --- //
const searchInput = document.getElementById("search");
const searchResults = document.getElementById("search-results");

function positionResults() {
  // 입력창 좌표 기준으로 dropdown 위치 설정
  if (searchResults.classList.contains("hidden")) return;
  const r = searchInput.getBoundingClientRect();
  searchResults.style.top = (r.bottom + 8) + "px";
  searchResults.style.left = r.left + "px";
  searchResults.style.width = r.width + "px";
}

function showResults() {
  searchResults.classList.remove("hidden");
  positionResults();
}
function hideResults() {
  searchResults.classList.add("hidden");
}

// 스크롤·리사이즈 시 dropdown 위치 추적
window.addEventListener("scroll", positionResults, { passive: true });
window.addEventListener("resize", positionResults);

function renderSearch() {
  const q = searchInput.value.trim();
  if (!q) { hideResults(); searchResults.innerHTML = ""; return; }
  const matches = DATA.filter(d => d.n && d.n.includes(q)).slice(0, 30);
  if (!matches.length) {
    searchResults.innerHTML = `<div class="search-empty"><span class="em">🔍</span>일치하는 후보자가 없습니다</div>`;
    showResults();
    return;
  }
  searchResults.innerHTML = matches.map(r => {
    const thumbHtml = safeImg(
      r.ph ? thumbUrl(r.ph) : "", "thumb", "",
      `<div class="thumb"></div>`
    );
    return `
    <div class="search-result" data-hubo="${r.h}">
      ${thumbHtml}
      <div class="meta">
        <div class="name-line">
          <span class="name">${r.n}</span>
          ${r.hj ? `<span class="hanja-mini">${r.hj}</span>` : ''}
        </div>
        <div class="sub">${r.e}${r.s ? ' · ' + r.s : ''}${r.g ? ' · ' + r.g : ''}</div>
      </div>
      <span class="party-pill" style="${pillStyle(r.p)}">${r.p}</span>
    </div>`;
  }).join("");
  showResults();
  for (const el of searchResults.querySelectorAll(".search-result")) {
    el.addEventListener("click", () => {
      openDetail(el.dataset.hubo);
      searchInput.value = "";
      hideResults();
    });
  }
}
let searchTimer;
searchInput.addEventListener("input", () => {
  clearTimeout(searchTimer);
  searchTimer = setTimeout(renderSearch, 120);
});
document.addEventListener("click", (e) => {
  // 검색창/결과 둘 다 바깥일 때만 닫기
  if (!e.target.closest(".search-wrap") && !e.target.closest("#search-results")) {
    hideResults();
  }
});
searchInput.addEventListener("focus", () => { if (searchInput.value.trim()) renderSearch(); });

// --- comparison modal --- //
const compareOverlay = document.getElementById("compare-overlay");
document.getElementById("compare-close").addEventListener("click", () => compareOverlay.classList.add("hidden"));
compareOverlay.addEventListener("click", (e) => { if (e.target === compareOverlay) compareOverlay.classList.add("hidden"); });

function openCompareModal() {
  const e = fEl.value, s = fSi.value, g = fSg.value;
  if (!e || !s) return;
  const d = currentFilter();
  document.getElementById("compare-eyebrow").textContent = e;
  let title = s;
  if (g) title += " · " + g;
  document.getElementById("compare-title").textContent = title;
  document.getElementById("compare-count").textContent = `${d.length}명`;

  const grid = document.getElementById("compare-grid");
  if (!d.length) {
    grid.innerHTML = `<div class="empty-state"><span class="em">📭</span>해당 조건의 후보자가 없습니다</div>`;
  } else {
    const sorted = [...d].sort((a,b) => {
      const ga = parseInt(a.gh) || 999, gb = parseInt(b.gh) || 999;
      return ga - gb;
    });
    grid.innerHTML = sorted.map(r => {
      const photoUrl = r.ph ? thumbUrl(r.ph) : "";
      const photoHtml = safeImg(photoUrl, "", r.n, `<span>사진 없음</span>`);
      const color = colorFor(r.p);
      const crimCls = r.c > 0 ? 'bad' : 'good';
      return `
        <div class="compare-card" data-hubo="${r.h}">
          <div class="strip" style="background:${color}"></div>
          <div class="photo-wrap">${photoHtml}</div>
          <div class="info">
            <div class="gh-party">
              ${r.gh ? `<span class="gh-num">${r.gh}</span>` : ''}
              <span style="${pillStyle(r.p)}padding:2px 8px;border-radius:999px;font-weight:600;">${r.p}</span>
            </div>
            <div class="name">${r.n}${r.hj ? ` <span class="hanja">(${r.hj})</span>` : ''}</div>
            <div class="stats">
              <div class="stat"><div class="k">나이</div><div class="v">${r.a || '-'}세</div></div>
              <div class="stat"><div class="k">자산</div><div class="v">${fmtAsset(r.as)}</div></div>
              <div class="stat"><div class="k">전과</div><div class="v ${crimCls}">${r.c > 0 ? r.c + '건' : '없음'}</div></div>
              <div class="stat"><div class="k">병역</div><div class="v">${r.gd === '남' ? r.m : '-'}</div></div>
            </div>
          </div>
        </div>`;
    }).join("");
    for (const card of grid.querySelectorAll(".compare-card")) {
      card.addEventListener("click", () => { compareOverlay.classList.add("hidden"); openDetail(card.dataset.hubo); });
    }
  }
  compareOverlay.classList.remove("hidden");
}

// --- main render --- //
function render() {
  destroyAll();
  const d = currentFilter();
  document.getElementById("count-display").innerHTML = `필터 결과 <strong>${d.length.toLocaleString()}명</strong>`;

  // KPI
  document.getElementById("kpi-count").textContent = d.length.toLocaleString();
  document.getElementById("kpi-count-sub").textContent = `전체 ${DATA.length.toLocaleString()}명 중`;

  const ages = d.map(x => x.a).filter(x => x);
  const avgAge = avg(ages);
  document.getElementById("kpi-age").textContent = avgAge !== null ? avgAge.toFixed(1) : "-";
  document.getElementById("kpi-age-sub").textContent = `중간값 ${median(ages) || "-"}세`;

  const assets = d.map(x => x.as).filter(x => x !== null && x !== undefined);
  const avgAs = avg(assets);
  document.getElementById("kpi-assets").innerHTML = `${fmtAsset(avgAs)}`;
  document.getElementById("kpi-assets-sub").textContent = `중간값 ${fmtAsset(median(assets))} · 신고 ${assets.length}/${d.length}명`;

  const crimN = d.filter(x => x.c > 0).length;
  document.getElementById("kpi-crim").textContent = d.length ? (crimN / d.length * 100).toFixed(1) : "-";
  document.getElementById("kpi-crim-sub").textContent = `${crimN.toLocaleString()}명 · 총 ${d.reduce((s,x)=>s+x.c,0).toLocaleString()}건`;

  const males = d.filter(x => x.gd === "남");
  const milMiss = males.filter(x => x.m === "미필" || x.m === "면제").length;
  document.getElementById("kpi-mil").textContent = males.length ? (milMiss / males.length * 100).toFixed(1) : "-";

  // 정당별 분포 (전체 탭일 때만 의미있음 — 특정 정당 탭일 때는 1개라 숨김)
  const sectionParty = document.getElementById("sec-party");
  if (activePartyTab === "전체") {
    sectionParty.style.display = "";
    const partyCount = {};
    for (const x of d) partyCount[x.p] = (partyCount[x.p] || 0) + 1;
    const partySorted = Object.entries(partyCount).sort((a,b) => b[1]-a[1]).slice(0, 15);
    makeBar("ch-party", partySorted.map(x=>x[0]), partySorted.map(x=>x[1]), partySorted.map(x=>colorFor(x[0])), { unit: "명" });
  } else {
    sectionParty.style.display = "none";
  }

  // ===== 정당별 비교 차트는 항상 DATA 전체로 계산 (필터 무관) =====
  // 이유: 정당 간 절대 비교를 유지해 필터 변경 시 막대가 출렁이지 않게 함
  const byParty = {};
  for (const x of DATA) {
    if (!byParty[x.p]) byParty[x.p] = { total: 0, crim: 0 };
    byParty[x.p].total++;
    if (x.c > 0) byParty[x.p].crim++;
  }
  const crimPartyEntries = Object.entries(byParty).filter(([,v])=>v.total >= 5)
    .sort((a,b) => (b[1].crim/b[1].total) - (a[1].crim/a[1].total)).slice(0, 12);
  makeBar("ch-crim-party",
    crimPartyEntries.map(([p,v]) => `${p} (${v.crim}/${v.total})`),
    crimPartyEntries.map(([,v]) => +(v.crim/v.total*100).toFixed(1)),
    crimPartyEntries.map(([p,]) => colorFor(p)),
    { unit: "%", tooltipFmt: (v) => v.toFixed(1) + "% (보유율)" });

  // 전과: 지역별
  const bySido = {};
  for (const x of d) {
    if (!x.s) continue;
    if (!bySido[x.s]) bySido[x.s] = { total: 0, crim: 0 };
    bySido[x.s].total++;
    if (x.c > 0) bySido[x.s].crim++;
  }
  const sidoEntries = Object.entries(bySido).sort((a,b)=> (b[1].crim/b[1].total) - (a[1].crim/a[1].total));
  makeBar("ch-crim-sido",
    sidoEntries.map(([s,v]) => `${s} (${v.crim}/${v.total})`),
    sidoEntries.map(([,v]) => +(v.crim/v.total*100).toFixed(1)),
    sidoEntries.map(() => '#5645d4'),
    { unit: "%", tooltipFmt: (v) => v.toFixed(1) + "%" });

  // 전과 Top 30
  const crimTop = [...d].filter(x=>x.c>0).sort((a,b) => b.c - a.c).slice(0, 30);
  fillTable("tbl-crim-top", crimTop, (r,i) =>
    `<td>${i+1}</td><td><strong>${r.n}</strong></td><td><span class="party-pill" style="${pillStyle(r.p)}">${r.p}</span></td><td>${r.s}${r.g?' · '+r.g:''}</td><td>${r.e}</td><td class="num">${r.c}건</td>`);

  // 자산: 정당별 (전체 기준 · 필터 무관)
  const assetByParty = {};
  for (const x of DATA) {
    if (x.as === null || x.as === undefined) continue;
    (assetByParty[x.p] = assetByParty[x.p] || []).push(x.as);
  }
  const assetPartyEntries = Object.entries(assetByParty).filter(([,arr])=>arr.length >= 5)
    .map(([p,arr]) => [p, avg(arr)]).sort((a,b)=>b[1]-a[1]).slice(0, 12);
  makeBar("ch-asset-party",
    assetPartyEntries.map(x=>x[0]),
    assetPartyEntries.map(x=>Math.round(x[1])),
    assetPartyEntries.map(x=>colorFor(x[0])),
    { unit: "원", tooltipFmt: (v) => fmtAsset(v), valueTickFmt: (v) => fmtAsset(v) });

  // 자산: 지역별
  const assetBySido = {};
  for (const x of d) {
    if (x.as === null || x.as === undefined || !x.s) continue;
    (assetBySido[x.s] = assetBySido[x.s] || []).push(x.as);
  }
  const assetSidoEntries = Object.entries(assetBySido).map(([s,arr]) => [s, avg(arr)]).sort((a,b)=>b[1]-a[1]);
  makeBar("ch-asset-sido",
    assetSidoEntries.map(x=>x[0]),
    assetSidoEntries.map(x=>Math.round(x[1])),
    assetSidoEntries.map(() => '#5645d4'),
    { unit: "원", tooltipFmt: (v) => fmtAsset(v), valueTickFmt: (v) => fmtAsset(v) });

  // 자산 히스토그램
  const buckets = [
    { label: "빚(음수)", lo: -Infinity, hi: 0 },
    { label: "0~1억", lo: 0, hi: 100000 },
    { label: "1~5억", lo: 100000, hi: 500000 },
    { label: "5~10억", lo: 500000, hi: 1000000 },
    { label: "10~30억", lo: 1000000, hi: 3000000 },
    { label: "30~50억", lo: 3000000, hi: 5000000 },
    { label: "50~100억", lo: 5000000, hi: 10000000 },
    { label: "100억+", lo: 10000000, hi: Infinity },
  ];
  const hist = buckets.map(b => d.filter(x => x.as !== null && x.as !== undefined && x.as >= b.lo && x.as < b.hi).length);
  makeBar("ch-asset-hist", buckets.map(b=>b.label), hist, buckets.map((_,i)=> i === 0 ? '#dc2626' : '#5645d4'), { horizontal: false, unit: "명" });

  // 부자/가난 Top
  const ranked = d.filter(x => x.as !== null && x.as !== undefined).sort((a,b) => b.as - a.as);
  const renderAssetRow = (r,i) =>
    `<td>${i+1}</td><td><strong>${r.n}</strong></td><td><span class="party-pill" style="${pillStyle(r.p)}">${r.p}</span></td><td>${r.s}${r.g?' · '+r.g:''}</td><td class="num" title="${(r.as*1000).toLocaleString()}원">${fmtAsset(r.as)}</td>`;
  fillTable("tbl-rich", ranked.slice(0, 30), renderAssetRow);
  fillTable("tbl-poor", ranked.slice(-30).reverse(), renderAssetRow);

  // 나이: 정당별 (전체 기준 · 필터 무관)
  const ageByParty = {};
  for (const x of DATA) { if (x.a) (ageByParty[x.p] = ageByParty[x.p] || []).push(x.a); }
  const agePartyEntries = Object.entries(ageByParty).filter(([,arr])=>arr.length >= 5)
    .map(([p,arr]) => [p, avg(arr)]).sort((a,b)=>b[1]-a[1]).slice(0, 12);
  makeBar("ch-age-party",
    agePartyEntries.map(x=>`${x[0]}`),
    agePartyEntries.map(x=>+x[1].toFixed(1)),
    agePartyEntries.map(x=>colorFor(x[0])),
    { unit: "세", tooltipFmt: (v) => v.toFixed(1) + "세" });

  // 나이: 지역별
  const ageBySido = {};
  for (const x of d) { if (x.a && x.s) (ageBySido[x.s] = ageBySido[x.s] || []).push(x.a); }
  const ageSidoEntries = Object.entries(ageBySido).map(([s,arr]) => [s, avg(arr)]).sort((a,b)=>b[1]-a[1]);
  makeBar("ch-age-sido",
    ageSidoEntries.map(x=>x[0]),
    ageSidoEntries.map(x=>+x[1].toFixed(1)),
    ageSidoEntries.map(() => '#5645d4'),
    { unit: "세", tooltipFmt: (v) => v.toFixed(1) + "세" });

  // 나이 히스토그램
  const ageBuckets = ["20대","30대","40대","50대","60대","70대 이상"];
  const ageHist = ageBuckets.map((_, i) => {
    const lo = 20 + i*10, hi = i === 5 ? Infinity : 30 + i*10;
    return d.filter(x => x.a !== null && x.a >= lo && x.a < hi).length;
  });
  makeBar("ch-age-hist", ageBuckets, ageHist, ageBuckets.map(()=>'#5645d4'), { horizontal: false, unit: "명" });

  // 병역
  const milCount = {};
  for (const x of d) { if (x.gd === "남") milCount[x.m] = (milCount[x.m] || 0) + 1; }
  // 보라 톤 위주 + 부정적(미필/면제) 빨강 강조
  const MIL_COLORS = { "복무필": "#5645d4", "미필": "#dc2626", "면제": "#e03131", "복무중": "#7b3ff2", "비대상": "#d6b6f6", "병적없음": "#a4a097", "기타": "#787671", "해당없음": "#d6b6f6" };
  const milLabels = Object.keys(milCount);
  makeDoughnut("ch-mil", milLabels, milLabels.map(k=>milCount[k]), milLabels.map(k=>MIL_COLORS[k] || "#a4a097"));

  // 직업
  const jobCount = {};
  for (const x of d) jobCount[x.j] = (jobCount[x.j] || 0) + 1;
  const jobEntries = Object.entries(jobCount).sort((a,b)=>b[1]-a[1]).slice(0, 14);
  makeBar("ch-job", jobEntries.map(x=>x[0]), jobEntries.map(x=>x[1]), jobEntries.map(()=>'#5645d4'), { unit: "명" });
}

render();

// =========== Floating Chatbot ===========
(function(){
  const fab = document.getElementById('chatFab');
  const panel = document.getElementById('chatPanel');
  const closeBtn = document.getElementById('chatClose');
  const form = document.getElementById('chatForm');
  const input = document.getElementById('chatInput');
  const sendBtn = document.getElementById('chatSend');
  const msgs = document.getElementById('chatMsgs');
  const suggestionsEl = document.getElementById('chatSuggestions');

  if (!fab || !panel) return;

  function openPanel(){
    panel.classList.remove('hidden');
    fab.classList.add('active');
    setTimeout(() => input.focus(), 60);
  }
  function closePanel(){
    panel.classList.add('hidden');
    fab.classList.remove('active');
  }
  fab.addEventListener('click', openPanel);
  closeBtn.addEventListener('click', closePanel);

  // 추천 칩 — 실제 후보를 한 명 골라서 끼워넣기
  const sampleCand = DATA.find(d => d.pl && d.pl.length && d.n) || null;
  const chips = sampleCand
    ? [
        `${sampleCand.n} 후보의 핵심 공약 알려줘`,
        `${sampleCand.n} 청년·일자리 공약은?`,
        `${sampleCand.n} 주거·교통 공약은?`,
      ]
    : ['후보 이름을 함께 입력해 주세요'];
  suggestionsEl.innerHTML = chips.map(s =>
    `<button type="button" class="qa-chip">${s}</button>`
  ).join('');
  suggestionsEl.querySelectorAll('.qa-chip').forEach(chip => {
    chip.addEventListener('click', () => ask(chip.textContent));
  });

  function addMsg(role, text, extraClass=''){
    const div = document.createElement('div');
    div.className = 'qa-msg ' + role + (extraClass ? ' '+extraClass : '');
    div.textContent = text;
    msgs.appendChild(div);
    const body = document.querySelector('#chatPanel .chat-body');
    if (body) body.scrollTop = body.scrollHeight;
    return div;
  }

  // 후보 이름을 질문에서 추출 — 긴 이름·정당·선거구 매칭에 가산점
  function findCandidates(q){
    const text = q.trim();
    if (!text) return [];
    const matches = [];
    for (const d of DATA) {
      if (!d.n || d.n.length < 2) continue;
      if (text.includes(d.n)) {
        const score =
          d.n.length * 10
          + (d.g && text.includes(d.g) ? 30 : 0)
          + (d.s && text.includes(d.s) ? 20 : 0)
          + (d.p && text.includes(d.p) ? 15 : 0)
          + (d.e && text.includes(d.e) ? 10 : 0);
        matches.push({ d, score });
      }
    }
    matches.sort((a, b) => b.score - a.score);
    return matches.map(m => m.d);
  }

  async function ask(question){
    if (!question.trim()) return;
    addMsg('user', question);
    input.value = '';
    sendBtn.disabled = true;
    try {
      const matches = findCandidates(question);
      if (matches.length === 0) {
        addMsg('bot',
          '어느 후보의 공약이 궁금하신가요? 후보 이름과 함께 다시 물어봐 주세요.\n예: "정원오 후보 청년 공약은?"',
          'error');
        return;
      }

      // 동명이인 처리 — top1과 top2 점수가 같으면 후보 안내
      if (matches.length >= 2) {
        const top = matches[0].n, ambiguous = [];
        for (const m of matches) {
          if (m.n === top) ambiguous.push(m); else break;
        }
        if (ambiguous.length > 1) {
          const lines = ambiguous.slice(0, 5).map(m =>
            `· ${m.n} (${m.p}, ${m.e}${m.g ? ' · ' + m.g : ''})`
          ).join('\n');
          addMsg('bot',
            `같은 이름의 후보가 ${ambiguous.length}명 있어요. 선거구나 정당을 함께 적어 주세요:\n${lines}`,
            'error');
          return;
        }
      }

      const cand = matches[0];
      if (!cand.pl || !cand.pl.length) {
        addMsg('bot',
          `${cand.n} 후보(${cand.p}, ${cand.e}${cand.g ? ' · ' + cand.g : ''})는 NEC에 제출된 공약 자료가 없어 답변할 수 없어요.`,
          'error');
        return;
      }

      const loading = addMsg('bot', `${cand.n} 후보의 공약을 읽고 답변 작성 중…`, 'loading');
      const resp = await fetch('/api/qa', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({
          question,
          candidate: { huboId: cand.h, name: cand.n, party: cand.p, giho: cand.gh,
                        subSgName: cand.e, sgg: cand.g, sido: cand.s },
          pledgeText: '',
          pdfUrls: cand.pl
        })
      });
      const j = await resp.json().catch(() => ({ error: '응답 파싱 실패' }));
      loading.remove();
      if (!resp.ok) {
        addMsg('bot', j.error || ('서버 오류 ' + resp.status), 'error');
        return;
      }
      // 답변 앞에 후보자 식별 정보 한 줄
      const tag = `[${cand.n} · ${cand.p} · ${cand.e}${cand.g ? ' · ' + cand.g : ''}]`;
      addMsg('bot', tag + '\n\n' + ((j.answer || '').trim() || '(빈 응답)'));

    } catch(e) {
      addMsg('bot', '요청 실패: ' + (e.message || e), 'error');
    } finally {
      sendBtn.disabled = false;
    }
  }

  form.addEventListener('submit', e => {
    e.preventDefault();
    ask(input.value);
  });

  // ESC 로 닫기
  document.addEventListener('keydown', e => {
    if (e.key === 'Escape' && !panel.classList.contains('hidden')) closePanel();
  });
})();

</script>
</body>
</html>
"""


def main():
    records = load_records()
    print(f"레코드 {len(records)}건 로드")
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    html_out = TEMPLATE.replace("__DATA__", payload)
    OUT.write_text(html_out, encoding="utf-8")
    print(f"생성: {OUT} ({OUT.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
